import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.utils import timezone

def export_absent_students_to_excel(queryset, branch_name):
    """
    Bugun kelmagan o'quvchilar ro'yxatini Excel formatida qaytaradi.
    """
    today = timezone.localdate()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bugun Kelmaganlar"

    # Header Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid") # Red for absence
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ["№", "O'quvchi ismi-familiyasi", "Guruh", "Telefon"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20

    # Data
    for row_num, att in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = att.student.full_name if att.student else "Noma'lum"
        ws.cell(row=row_num, column=3).value = att.group.name if att.group else "Noma'lum"
        ws.cell(row=row_num, column=4).value = att.student.phone if att.student else ""

    from urllib.parse import quote
    raw_filename = f"kelmaganlar_{branch_name}_{today}.xlsx"
    # Remove characters that can break the HTTP header or represent directory traversal
    safe_filename = raw_filename.replace('\n', '').replace('\r', '').replace('"', '_').replace('/', '_').replace('\\', '_')
    encoded_filename = quote(safe_filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"; filename*=utf-8\'\'{encoded_filename}'
    wb.save(response)
    return response
