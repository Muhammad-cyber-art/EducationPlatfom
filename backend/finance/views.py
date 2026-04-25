from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status, permissions ,filters
from rest_framework.decorators import action ,api_view, permission_classes
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from django.utils import timezone
from rest_framework.views import APIView
from .models import Payment, EmployeePayment ,StaffProfile, FinanceTransaction, EmployeeAdvance
from .serializers import (
    PaymentSerializer, EmployeePaymentSerializer, 
    StaffProfileSerializer, FinanceTransactionSerializer,
    EmployeeAdvanceSerializer
)
from django.contrib.auth import get_user_model
from django.db.models import Sum, Q, F, Count ,OuterRef, Subquery
from groups.models import Group, Student, Branch
from homework_attends.models import Attendance
from rest_framework.permissions import IsAuthenticated
from .services import generate_monthly_payments
from permissions.permissions import HasModulePermission
User = get_user_model()
# 1. O'quvchilar to'lovlari ViewSet

class StudentPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'finance'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]

    # Filterlar
    filterset_fields = {
        'is_paid': ['exact'],
        'month': ['exact', 'gte', 'lte'],
        'group__mentor': ['exact'],
        'group': ['exact'],
        'student': ['exact'],
        'group__branch': ['exact'],
    }

    search_fields = [
        'student__full_name',
        'student__phone',
        'student__parent_name',
        'student__parent_phone',
    ]

    ordering_fields = ['month', 'amount', 'created_at', 'paid_at']
    ordering = ['-month']

    def get_queryset(self):
        user = self.request.user
        qs = Payment.objects.select_related(
            'student',
            'group',
            'group__branch',
            'group__mentor',
            'marked_by'
        )

        # 1. Ruxsatlar bo'yicha filter
        if user.role == 'super_admin':
            pass # Hammani ko'radi
        elif user.role == 'admin':
            # Admin o'z filialidagi o'quvchilarning to'lovlarini ko'ra olishi kerak.
            # Agar guruh o'chib ketgan bo'lsa (group=None), student__branch orqali tekshiramiz.
            qs = qs.filter(Q(group__branch=user.branch) | Q(student__branch=user.branch))
        elif user.role == 'mentor':
            qs = qs.filter(group__mentor=user)
        else:
            return Payment.objects.none()

        # 2. LIST uchun mantiq: Agar oy filtri berilmagan bo'lsa, default joriy oyni chiqarish
        # Bu o'quvchilar ro'yxatda 10 marta takrorlanib ketmasligi uchun kerak
        if self.action == 'list':
            month_param = self.request.query_params.get('month')
            month_gte = self.request.query_params.get('month__gte')
            student_param = self.request.query_params.get('student')
            
            if not month_param and not month_gte and not student_param:
                current_month = timezone.localdate().replace(day=1)
                qs = qs.filter(month=current_month)

        return qs

    def retrieve(self, request, *args, **kwargs):
        """
        Bitta to'lovga kirilganda o'sha o'quvchining shu guruhdagi 
        barcha oylik to'lovlar tarixini ham qo'shib qaytaradi.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)

        # 1. Oylik to'lovlar
        history_qs = Payment.objects.filter(
            student=instance.student,
            group=instance.group
        ).select_related('marked_by').order_by('-month')
        history_serializer = PaymentSerializer(history_qs, many=True)

        # 2. Qo'shimcha tranzaksiyalar (Custom payments)
        extra_qs = FinanceTransaction.objects.filter(
            student=instance.student,
            group=instance.group
        ).select_related('marked_by').order_by('-date')
        extra_serializer = FinanceTransactionSerializer(extra_qs, many=True)

        # Asosiy ma'lumotga tarixni qo'shamiz
        data = serializer.data
        data['payment_history'] = history_serializer.data
        data['extra_transactions'] = extra_serializer.data
        
        return Response(data)

    @action(detail=False, methods=['post'], url_path='custom-payment')
    def custom_payment(self, request):
        """O'quvchi uchun maxsus (oylikdan tashqari) to'lov qabul qilish"""
        from .serializers import CustomPaymentSerializer
        serializer = CustomPaymentSerializer(data=request.data)
        if serializer.is_valid():
            id_student = serializer.validated_data['student']
            id_group = serializer.validated_data['group']
            amount = serializer.validated_data['amount']
            transaction_type = serializer.validated_data['transaction_type']
            payer_name = serializer.validated_data.get('payer_name', '')
            description = serializer.validated_data.get('description', '')
            date_val = serializer.validated_data.get('date', timezone.localdate())
            
            student = get_object_or_404(Student, id=id_student)
            group = get_object_or_404(Group, id=id_group)
            
            # Filialni aniqlash (ustuvorlik: o'quvchi -> guruh -> admin)
            branch_instance = student.branch or group.branch
            if not branch_instance and hasattr(request.user, 'branch'):
                branch_instance = request.user.branch

            if not branch_instance:
                return Response({
                    "detail": "Filialni aniqlab bo'lmadi. O'quvchi yoki guruh birorta filialga biriktirilgan bo'lishi shart."
                }, status=status.HTTP_400_BAD_REQUEST)

            # To'lovchi ismi
            final_payer = payer_name if (payer_name and payer_name.strip()) else student.full_name

            # Sarlavha (Kirim yoki Chiqimga qarab)
            if transaction_type == 'income':
                title_val = f"Qo'shimcha to'lov: {student.full_name}"
            else:
                title_val = f"Chiqim (Portal): {student.full_name}"

            try:
                ft = FinanceTransaction.objects.create(
                    transaction_type=transaction_type,
                    category='student_extra',
                    amount=amount,
                    date=date_val,
                    marked_by=request.user,
                    branch=branch_instance,
                    student=student,
                    group=group,
                    payer_name=final_payer,
                    title=title_val,
                    description=description
                )
                
                # ✅ TUZATISH: Guruh mentorining oyligini avtomatik qayta hisoblash
                # Qo'shimcha to'lov mentor oyligiga ta'sir qilishi kerak (foiz hisoblash uchun)
                mentor_salary_updated = False
                try:
                    mentor = group.mentor
                    if mentor and hasattr(mentor, 'staff_profile'):
                        profile = mentor.staff_profile
                        if profile.salary_type in ['percentage', 'student_count']:
                            # O'sha oydagi to'lanmagan EmployeePayment ni topib yangilash
                            import calendar
                            from datetime import date
                            payment_month = date_val.replace(day=1)
                            
                            emp_payment = EmployeePayment.objects.filter(
                                employee=mentor,
                                month=payment_month,
                                is_paid=False  # Faqat to'lanmagan bo'lsa yangilaymiz
                            ).first()
                            
                            if emp_payment:
                                new_salary = profile.calculate_salary_for_month(payment_month)
                                emp_payment.salary_base = new_salary
                                emp_payment.save()
                                mentor_salary_updated = True
                            else:
                                # Agar payment mavjud bo'lmasa — yangi yaratib qo'yamiz
                                new_salary = profile.calculate_salary_for_month(payment_month)
                                EmployeePayment.objects.get_or_create(
                                    employee=mentor,
                                    month=payment_month,
                                    defaults={
                                        'salary_base': new_salary,
                                        'bonus': 0,
                                        'deductions': 0,
                                        'is_paid': False
                                    }
                                )
                                mentor_salary_updated = True
                except Exception as salary_err:
                    # Oylik hisoblashda xato bo'lsa, asosiy to'lovni bekor qilmaymiz
                    import traceback
                    print(f"Mentor salary recalculate error: {salary_err}")
                    traceback.print_exc()
                
                return Response({
                    "status": "Muvaffaqiyatli saqlandi", 
                    "id": ft.id,
                    "amount": str(ft.amount),
                    "student": student.full_name,
                    "mentor_salary_updated": mentor_salary_updated
                })
            except Exception as e:
                # 500 xato o'rniga aniqroq xabar qaytaramiz (audit uchun)
                return Response({
                    "detail": f"Tranzaksiyani saqlashda xatolik yuz berdi: {str(e)}"
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='confirm')
    def confirm(self, request, pk=None):
        """
        Istalgan ID dagi (joriy yoki tarixdagi) to'lovni tasdiqlash
        """
        payment = self.get_object()

        if payment.is_paid:
            return Response({"detail": "Bu to'lov allaqachon to'langan"}, status=400)

        if request.user.role not in ['admin', 'super_admin']:
            return Response({"detail": "Faqat adminlar to'lovni tasdiqlay oladi"}, status=403)

        # Refundni e'tiborsiz qoldirish xususiyati
        ignore_refund = request.data.get('ignore_refund', False)
        ignore_refund = str(ignore_refund).lower() in ['true', '1', 'yes']
        payment.refund_ignored = ignore_refund
        
        # To'lov summasini o'zgartirish (agar so'rovda yangi summa kelsa)
        new_amount = request.data.get('amount')
        if new_amount is not None:
            payment.amount = new_amount
            payment.save()
            
            # Kelgusi oylar uchun ham saqlab qolish (individual narx sifatida)
            student = payment.student
            student.custom_fee = new_amount
            
            # Agar status hali oddiy bo'lsa, uni "Kelishilgan narx"ga o'tkazib qo'yamiz
            if student.status == 'regular':
                student.status = 'negotiated'
            student.save()

        else:
            # Agar aniq summa kiritilmagan bo'lsa va refund ignored bo'lmasa, uni ayiramiz
            if not payment.refund_ignored:
                from decimal import Decimal
                from finance.utils import floor_amount
                refund = payment.student.calculate_refund_amount(payment.month.year, payment.month.month)
                if float(refund) > 0:
                    current_amount = Decimal(str(payment.amount or payment.group.monthly_price))
                    payment.amount = floor_amount(current_amount - Decimal(str(refund)))
            payment.save()


        # To'lovni tasdiqlash (Model metodidan foydalanamiz, shunda Ledgerga yozadi)
        payment.mark_as_paid(request.user)
        
        # ✅ TUZATISH: Student to'lovi tasdiqlanganida mentor oyligini qayta hisoblash
        try:
            mentor = payment.group.mentor
            if mentor and hasattr(mentor, 'staff_profile'):
                profile = mentor.staff_profile
                if profile.salary_type in ['percentage', 'student_count']:
                    payment_month = payment.month.replace(day=1) if payment.month else None
                    if payment_month:
                        emp_payment = EmployeePayment.objects.filter(
                            employee=mentor,
                            month=payment_month,
                            is_paid=False
                        ).first()
                        if emp_payment:
                            new_salary = profile.calculate_salary_for_month(payment_month)
                            emp_payment.salary_base = new_salary
                            emp_payment.save()
        except Exception as salary_err:
            import traceback
            print(f"[confirm] Mentor salary recalculate error: {salary_err}")
            traceback.print_exc()

        return Response({
            "status": "To'lov muvaffaqiyatli tasdiqlandi",
            "student": payment.student.full_name,
            "month": payment.month.strftime('%Y-%m') if payment.month else "Noma'lum",
            "amount": str(payment.amount)
        })

    def perform_update(self, serializer):
        """
        Update orqali is_paid o'zgartirilsa, marked_by va paid_at ni avtomatik yozish
        """
        instance = serializer.instance
        is_paid = serializer.validated_data.get('is_paid', instance.is_paid)
        
        if is_paid and not instance.is_paid:
            if self.request.user.role in ['admin', 'super_admin']:
                serializer.save(marked_by=self.request.user, paid_at=timezone.now())
            else:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("To'lovni tasdiqlash uchun ruxsatingiz yo'q.")
        else:
            serializer.save()
# 2. Xodimlar maoshlari ViewSet

class EmployeePaymentViewSet(viewsets.ModelViewSet):
    serializer_class = EmployeePaymentSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'finance'
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # 1. Filterlash maydonlarini aniq ko'rsatamiz
    filterset_fields = {
        'is_paid': ['exact'],
        'month': ['exact', 'gte', 'lte'],
        'employee__role': ['exact'],      # Role orqali filter (mentor, admin)
        'employee__branch': ['exact'],    # Filial orqali filter
    }
    
    search_fields = ['employee__first_name', 'employee__last_name', 'employee__username']
    ordering_fields = ['month', 'salary_base', 'created_at']
    ordering = ['-month']

    def get_queryset(self):
        user = self.request.user
        # Select related orqali DB ga so'rovlarni kamaytiramiz
        qs = EmployeePayment.objects.select_related(
            'employee', 
            'employee__branch', 
            'employee__staff_profile', 
            'marked_by'
        )

        # 2. Ruxsatlar (Role-based access)
        if user.role == 'super_admin':
            pass # Hamma narsani ko'radi
        elif user.role == 'admin':
            qs = qs.filter(employee__branch=user.branch)
        else:
            # Mentor yoki boshqalar faqat o'zinikini ko'radi
            qs = qs.filter(employee=user)

        # 3. List (Ro'yxat) uchun aqlli oylik filter
        if self.action == 'list':
            # Agar URL'da oy filtrlanmagan bo'lsa, avtomatik joriy oyni ko'rsatadi
            # Masalan: /api/finance/employee-payment/ -> joriy oy chiqadi
            # Masalan: /api/finance/employee-payment/?month=2024-12-01 -> dekabr chiqadi
            month_param = self.request.query_params.get('month')
            month_gte = self.request.query_params.get('month__gte')
            
            if not month_param and not month_gte:
                current_month = timezone.localdate().replace(day=1)
                qs = qs.filter(month=current_month)
        
        return qs
    @action(detail=False, methods=['get'])
    def current(self, request):
        """
        Tizimga kirgan xodimning joriy (oxirgi) maosh ma'lumotlarini 
        tarixi bilan birga qaytaradi.
        """
        instance = EmployeePayment.objects.filter(
            employee=request.user
        ).order_by('-month').first()
        
        if not instance:
            return Response({
                "detail": "Siz uchun hali to'lov ma'lumotlari generatsiya qilinmagan."
            }, status=404)
            
        serializer = self.get_serializer(instance)
        
        # Tarixni qo'shish
        history_qs = EmployeePayment.objects.filter(
            employee=instance.employee
        ).order_by('-month')
        
        history_serializer = EmployeePaymentSerializer(history_qs, many=True)
        
        data = serializer.data
        data['payment_history'] = history_serializer.data
        return Response(data)

    @action(detail=True, methods=['post'], url_path='confirm')
    def confirm(self, request, pk=None):
        """
        Istalgan ID dagi to'lovni tasdiqlash (Tarixdagi to'lovlar uchun ham)
        """
        try:
            payment = self.get_object()
        except Exception as e:
            return Response({
                "detail": f"To'lov topilmadi: {str(e)}"
            }, status=404)
        
        if payment.is_paid:
            return Response({"detail": "Bu maosh allaqachon to'langan"}, status=400)

        # Faqat super_admin tasdiqlay olsin (xavfsizlik uchun)
        if request.user.role != 'super_admin':
            return Response({"detail": "Faqat super_admin maoshni tasdiqlashi mumkin"}, status=403)

        # Maoshni tasdiqlash (Model metodidan foydalanamiz, shunda Ledgerga yozadi)
        try:
            # Yangi: Bonus va Ayirmalarni yangilash (agar kelgan bo'lsa)
            if 'bonus' in request.data:
                payment.bonus = request.data.get('bonus')
            if 'deductions' in request.data:
                payment.deductions = request.data.get('deductions')
                
            payment.mark_as_paid(request.user)
            
            return Response({
                "status": "Maosh to'lovi muvaffaqiyatli tasdiqlandi",
                "id": payment.id,
                "month": payment.month,
                "amount": str(payment.total_amount)
            })
        except Exception as e:
            # Xatolikni batafsil log qilish
            import traceback
            print(f"Error in confirm payment {pk}:")
            print(traceback.format_exc())
            
            return Response({
                "detail": f"To'lovni tasdiqlashda xatolik: {str(e)}"
            }, status=400)

     
    def retrieve(self, request, *args, **kwargs):
        """
        Bitta ma'lumotga kirganda (ID orqali), 
        o'sha xodimning barcha oylardagi to'lov tarixini ham qo'shib qaytaradi.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        
        # O'sha xodimning hamma vaqtdagi to'lovlari
        history_qs = EmployeePayment.objects.filter(
            employee=instance.employee
        ).order_by('-month')
        
        history_serializer = EmployeePaymentSerializer(history_qs, many=True)
        
        # Response ma'lumotlarini yig'amiz
        data = serializer.data
        data['payment_history'] = history_serializer.data
        return Response(data)

    def perform_update(self, serializer):
        """To'lov tasdiqlanganda kim va qachon tasdiqlaganini yozish"""
        instance = serializer.instance
        is_paid = serializer.validated_data.get('is_paid', instance.is_paid)
        
        if is_paid is True and instance.is_paid is False:
            serializer.save(marked_by=self.request.user, paid_at=timezone.now())
        else:
            serializer.save()

    @action(detail=True, methods=['post'], url_path='add-advance')
    def add_advance(self, request, pk=None):
        """Xodimga avans qo'shish"""
        payment = self.get_object()
        
        # Faqat super_admin avans bera olsin
        if request.user.role != 'super_admin':
            return Response({"detail": "Faqat super_admin avans bera oladi"}, status=403)
            
        amount = request.data.get('amount')
        description = request.data.get('description', '')
        date_val = request.data.get('date', timezone.now().date())
        
        if not amount or float(amount) <= 0:
            return Response({"detail": "Avans summasi noto'g'ri"}, status=400)
            
        try:
            from .models import EmployeeAdvance
            advance = EmployeeAdvance.objects.create(
                employee=payment.employee,
                month=payment.month, # Avans shu oy uchun
                amount=amount,
                description=description,
                date=date_val,
                marked_by=request.user
            )
            
            return Response({
                "success": True,
                "message": "Avans muvaffaqiyatli qo'shildi",
                "id": advance.id,
                "amount": float(advance.amount),
                "total_advances": float(payment.total_advances)
            })
        except Exception as e:
            return Response({"detail": f"Avans qo'shishda xatolik: {str(e)}"}, status=400)

    @action(detail=True, methods=['post'], url_path='recalculate')
    def recalculate(self, request, pk=None):
        """Maoshni o'sha oydagi tushumlarga qarab qayta hisoblash"""
        payment = self.get_object()
        
        if payment.is_paid:
            return Response({
                "error": "To'langan maoshni qayta hisoblab bo'lmaydi. Avval to'lovni bekor qiling."
            }, status=400)
        
        # Services.py dagi hisoblash funksiyasini chaqiramiz
        from .services import calculate_employee_salary
        profile = payment.employee.staff_profile
        
        new_salary = calculate_employee_salary(profile, payment.month)
        
        payment.salary_base = new_salary
        payment.save()
        
        return Response({
            "success": True,
            "message": "Maosh qayta hisoblandi",
            "new_salary": float(new_salary),
            "month": payment.month
        })

class StaffProfileViewSet(viewsets.ModelViewSet):
    serializer_class = StaffProfileSerializer
    
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = {
        'user': ['exact'],
        'user__role': ['exact'],
        'user__branch': ['exact'],
        'salary_type': ['exact'],
    }
    
    search_fields = ['user__first_name', 'user__last_name', 'user__username', 'karta']
    ordering_fields = ['fixed_salary', 'commission_percentage', 'created_at']
    ordering = ['-created_at']
    lookup_field = 'user'
    module_name = 'finance'

    def get_queryset(self):
        return StaffProfile.objects.select_related('user', 'user__branch').all()
    
    def get_permissions(self):
        return [permissions.IsAuthenticated(), HasModulePermission()]

    @action(detail=False, methods=['get'], url_path='available-staff')
    def available_staff(self, request):
        """Profili yo'q xodimlarni ko'rsatish"""
        staff_with_profile = StaffProfile.objects.values_list('user_id', flat=True)
        branch_id = request.query_params.get('branch')
        
        available_users = User.objects.filter(
            role__in=['mentor', 'admin'],
            is_active=True
        ).exclude(id__in=staff_with_profile)

        if branch_id:
            available_users = available_users.filter(branch_id=branch_id)
        
        data = [{
            "id": u.id, 
            "first_name": u.first_name,
            "last_name": u.last_name,
            "role": u.role
        } for u in available_users]
        
        return Response(data)
    
    @action(detail=True, methods=['post'], url_path='calculate-salary')
    def calculate_salary(self, request, user=None):
        """Xodimning maoshini hisoblash"""
        profile = self.get_object()
        month_str = request.data.get('month')
        
        if not month_str:
            return Response(
                {"error": "month parametri kerak (format: YYYY-MM-DD)"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from datetime import datetime
        month = datetime.strptime(month_str, "%Y-%m-%d").date()
        salary = profile.calculate_salary_for_month(month)
        
        return Response({
            "user": profile.user.get_full_name() or profile.user.username,
            "salary_type": profile.salary_type,
            "month": month_str,
            "calculated_salary": float(salary)
        })
    
    @action(detail=True, methods=['post'], url_path='recalculate-payment')
    def recalculate_payment(self, request, user=None):
        """Xodimning paymentini yaratish/yangilash"""
        profile = self.get_object()
        
        # Oyni olish
        month_str = request.data.get('month')
        if month_str:
            from datetime import datetime
            month = datetime.strptime(month_str, "%Y-%m-%d").date()
        else:
            month = timezone.now().date().replace(day=1)
        
        # Maoshni hisoblash
        calculated_salary = profile.calculate_salary_for_month(month)
        
        # EmployeePayment yaratish/yangilash
        emp_payment, created = EmployeePayment.objects.update_or_create(
            employee=profile.user,
            month=month,
            defaults={
                'salary_base': calculated_salary,
                'bonus': 0,
                'deductions': 0,
                'is_paid': False
            }
        )
        
        # Qo'shimcha ma'lumot
        extra_data = {}
        if profile.salary_type == 'percentage':
            total_income = profile.calculate_monthly_income(month)
            extra_data = {
                "total_income": float(total_income),
                "commission_percentage": float(profile.commission_percentage)
            }
        
        return Response({
            "success": True,
            "created": created,
            "user": profile.user.get_full_name() or profile.user.username,
            "salary_type": profile.salary_type,
            "month": month.strftime("%Y-%m"),
            "calculated_salary": float(calculated_salary),
            "employee_payment_id": emp_payment.id,
            **extra_data
        })
    
    @action(detail=True, methods=['get'], url_path='mentor-groups')
    def mentor_groups(self, request, user=None):
        """Mentorning guruhlarini ko'rish"""
        profile = self.get_object()
        
        if profile.user.role != 'mentor':
            return Response({
                "error": "Bu foydalanuvchi mentor emas"
            }, status=400)
        
        from groups.models import Group
        mentor_groups = Group.objects.filter(
            mentor=profile.user,
            is_faol=True
        ).select_related('branch')
        
        data = [{
            "id": g.id,
            "name": g.name,
            "monthly_price": float(g.monthly_price),
            "branch": g.branch.name if g.branch else None,
            "students_count": g.students.count()
        } for g in mentor_groups]
        
        return Response({
            "mentor": profile.user.get_full_name() or profile.user.username,
            "total_groups": mentor_groups.count(),
            "groups": data
        })
    
    @action(detail=True, methods=['get'], url_path='mentor-income')
    def mentor_income(self, request, user=None):
        """Mentorning oylik daromadini hisoblash"""
        profile = self.get_object()
        
        if profile.user.role != 'mentor':
            return Response({
                "error": "Bu foydalanuvchi mentor emas"
            }, status=400)
        
        # Oyni olish
        month_str = request.query_params.get('month')
        if month_str:
            from datetime import datetime
            month = datetime.strptime(month_str, "%Y-%m-%d").date()
        else:
            month = timezone.now().date().replace(day=1)
        
        # Guruhlar
        from groups.models import Group
        mentor_groups = Group.objects.filter(
            mentor=profile.user,
            is_faol=True
        ).select_related('branch')
        
        # Har bir guruhning ma'lumotlari
        total_expected = 0
        total_paid = 0
        groups_data = []
        
        for group in mentor_groups:
            students_count = group.students.count()
            expected = float(group.monthly_price or 0) * students_count
            
            # To'langan to'lovlar
            paid = Payment.objects.filter(
                group=group,
                month__year=month.year,
                month__month=month.month,
                is_paid=True
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            total_expected += expected
            total_paid += float(paid)
            
            groups_data.append({
                'id': group.id,
                'name': group.name,
                'monthly_price': float(group.monthly_price or 0),
                'branch_name': group.branch.name if group.branch else None,
                'students_count': students_count,
                'expected_income': expected,
                'paid_income': float(paid)
            })
        
        # Komissiya hisoblash
        commission = total_paid * (float(profile.commission_percentage) / 100)
        
        return Response({
            "mentor": profile.user.get_full_name() or profile.user.username,
            "month": month.strftime("%Y-%m"),
            "salary_type": profile.salary_type,
            "commission_percentage": float(profile.commission_percentage),
            "total_groups": mentor_groups.count(),
            "expected_income": total_expected,
            "paid_income": total_paid,
            "calculated_salary": commission,
            "groups": groups_data
        })
    

class FinanceDashboardView(APIView):
    # Oldin: permission_classes = [permissions.IsAuthenticated]
    # Endi: HasModulePermission ni qo'shamiz va modul nomini belgilaymiz
    permission_classes = [permissions.IsAuthenticated, HasModulePermission] 
    module_name = 'finance' # Bu view 'finance' moduliga tegishli ekanligini bildiradi

    def get(self, request):
        # 1. Sana parametrlarini olish
        today = timezone.localdate()
        month = int(request.query_params.get('month', today.month))
        year = int(request.query_params.get('year', today.year))
        
        user = request.user
        
        # Boshlang'ich filterlar
        payment_filter = Q(month__month=month, month__year=year)
        employee_filter = Q(month__month=month, month__year=year)
        
        # Ruxsat tekshirish
        if user.role == 'admin':
            payment_filter &= Q(group__branch=user.branch)
            employee_filter &= Q(employee__branch=user.branch)
        elif user.role != 'super_admin':
            return Response({"error": "Ruxsat yo'q"}, status=403)

        # 2. Umumiy daromad (Hamma tushumlar - Centralized Ledger dan)
        transaction_filter = Q(date__month=month, date__year=year)
        if user.role == 'admin':
            transaction_filter &= Q(branch=user.branch)
            
        total_income = FinanceTransaction.objects.filter(
            transaction_filter,
            transaction_type='income'
        ).aggregate(total=Sum('amount'))['total'] or 0

        # 3. Umumiy xarajat (Hamma chiqimlar - Centralized Ledger dan)
        total_expense = FinanceTransaction.objects.filter(
            transaction_filter,
            transaction_type='expense'
        ).aggregate(total=Sum('amount'))['total'] or 0

        # 4. Qarzdorlik (Hali to'lanmagan o'quvchi to'lovlari)
        total_debt = Payment.objects.filter(
            payment_filter, 
            is_paid=False
        ).aggregate(
            total=Sum('amount')
        )['total'] or 0

        # 5. Filiallar bo'yicha statistika
        branches_data = []
        if user.role == 'super_admin':
            branches = Branch.objects.all()
            for branch in branches:
                # Filial daromadi (Ledger dan)
                b_income = FinanceTransaction.objects.filter(
                    branch=branch,
                    date__month=month,
                    date__year=year,
                    transaction_type='income'
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                # Filial xarajati (Ledger dan)
                b_expense = FinanceTransaction.objects.filter(
                    branch=branch,
                    date__month=month,
                    date__year=year,
                    transaction_type='expense'
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                branches_data.append({
                    "id": branch.id,
                    "name": branch.name,
                    "income": float(b_income),
                    "expense": float(b_expense),
                    "profit": float(b_income - b_expense)
                })

        # 6. Top Guruhlar
        top_groups_qs = Group.objects.filter(is_faol=True)
        if user.role == 'admin':
            top_groups_qs = top_groups_qs.filter(branch=user.branch)
            
        top_groups_data = top_groups_qs.annotate(
            group_income=Sum(
                'payments__amount', 
                filter=Q(
                    payments__month__month=month, 
                    payments__month__year=year,
                    payments__is_paid=True
                )
            ),
            st_count=Count('students', distinct=True)
        ).order_by('-group_income')[:5]

        top_groups = [
            {
                "name": g.name,
                "profit": float(g.group_income or 0),
                "student_count": g.st_count
            } for g in top_groups_data
        ]

        # 7. O'tgan oy ma'lumotlari (Trendlar uchun)
        prev_month = month - 1 if month > 1 else 12
        prev_year = year if month > 1 else year - 1
        
        prev_tx_filter = Q(date__month=prev_month, date__year=prev_year)
        if user.role == 'admin':
            prev_tx_filter &= Q(branch=user.branch)

        prev_income = FinanceTransaction.objects.filter(prev_tx_filter, transaction_type='income').aggregate(total=Sum('amount'))['total'] or 0
        prev_expense = FinanceTransaction.objects.filter(prev_tx_filter, transaction_type='expense').aggregate(total=Sum('amount'))['total'] or 0

        # Trendlar (%)
        def calculate_trend(current, previous):
            curr = float(current)
            prev = float(previous)
            if prev == 0: 
                return 100.0 if curr > 0 else (-100.0 if curr < 0 else 0.0)
            return round(((curr - prev) / abs(prev)) * 100, 1)

        income_trend = calculate_trend(total_income, prev_income)
        expense_trend = calculate_trend(total_expense, prev_expense)
        profit_trend = calculate_trend(total_income - total_expense, prev_income - prev_expense)

        # 8. Yangi to'lovlar (Bugun qabul qilinganlar) - Badge uchun
        new_payments_count = Payment.objects.filter(paid_at__date=today, is_paid=True).count()

        # 9. Collection Efficiency (%)
        efficiency = 0
        if (total_income + total_debt) > 0:
            efficiency = round((total_income / (total_income + total_debt)) * 100, 1)

        # 10. Global Statistics (General counts)
        # Faqat faol guruhlarga ega o'quvchilarni sanaymiz
        total_students = Student.objects.filter(enrollments__is_active=True).distinct().count()
        total_mentors = User.objects.filter(role='mentor', is_active=True).count()
        total_groups = Group.objects.filter(is_faol=True).count()
        total_admins = User.objects.filter(role='admin', is_active=True).count()
        
        # Today's Absents
        today_absents = Attendance.objects.filter(
            date=today,
            is_present=False
        ).count()

        data = {
            "total_income": float(total_income),
            "total_expense": float(total_expense),
            "net_profit": float(total_income - total_expense),
            "total_debt": float(total_debt),
            "income_trend": income_trend,
            "expense_trend": expense_trend,
            "profit_trend": profit_trend,
            "collection_efficiency": efficiency,
            "new_payments_count": new_payments_count,
            "branches": branches_data,
            "top_groups": top_groups,
            "groups": top_groups,
            "stats": {
                "students": total_students,
                "mentors": total_mentors,
                "groups": total_groups,
                "admins": total_admins,
                "attendance_today": {
                    "absent": today_absents
                }
            },
            "finance": {
                "received_income": float(total_income),
                "expenses": float(total_expense),
                "net_profit": float(total_income - total_expense)
            }
        }

        return Response(data)

@api_view(['POST'])  # Faqat POST so'rovga javob beradi
@permission_classes([IsAuthenticated])  # Faqat login qilgan userlar
def trigger_monthly_payments(request):
    """Qo'lda to'lov varaqalarini yaratish"""
    try:
        # generate_monthly_payments() funksiyasini chaqiradi
        count = generate_monthly_payments()
        
        # Muvaffaqiyatli javob
        return Response({
            'success': True,
            'message': f'Muvaffaqiyatli: {count} ta varaqa yaratildi.',
            'count': count
        }, status=200)
        
    except Exception as e:
        # Xatolik bo'lsa
        return Response({
            'success': False,
            'message': f'Xatolik: {str(e)}'
        }, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def trigger_absence_refunds(request):
    """Qo'lda dars qoldirganlik uchun refundlarni hisoblash"""
    try:
        from .services import process_absence_refunds
        month_str = request.data.get('month')
        if month_str:
            from datetime import datetime
            month_date = datetime.strptime(month_str, "%Y-%m-%d").date()
        else:
            month_date = timezone.localdate().replace(day=1)
            
        count, amount = process_absence_refunds(month_date)
        
        return Response({
            'success': True,
            'message': f'Muvaffaqiyatli: {count} ta refund yaratildi. Jami summa: {amount}',
            'count': count,
            'amount': float(amount)
        }, status=200)
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Xatolik: {str(e)}'
        }, status=500)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from django.db.models import Q, Sum, Count, F
from django.utils import timezone

class BranchFinanceDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, branch_id):
        from groups.models import Student, Group, WaitingStudent
        from branches.models import Branch
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = request.user
        
        # Ruxsat tekshirish
        if user.role == 'admin':
            has_access = False
            # 1. Asosiy filial
            if user.branch_id == branch_id:
                has_access = True
            # 2. Qo'shimcha ruxsatlar
            elif hasattr(user, 'branch_accesses') and user.branch_accesses.filter(branch_id=branch_id).exists():
                has_access = True
            
            if not has_access:
                return Response({"error": "Bu filial statistikasini ko'rishga ruxsatingiz yo'q"}, status=403)
        
        elif user.role != 'super_admin':
            return Response({"error": "Ruxsat yo'q"}, status=403)

        # Sana parametrlari
        today = timezone.localdate()
        month = int(request.query_params.get('month', today.month))
        year = int(request.query_params.get('year', today.year))

        try:
            branch = Branch.objects.get(id=branch_id)
        except Branch.DoesNotExist:
            return Response({"error": "Filial topilmadi"}, status=404)

        # 1. Asosiy hisoblar
        # Mentorlar soni
        mentors_count = User.objects.filter(branch=branch, role='mentor', is_active=True).count()
        
        # Adminlar soni
        admins_count = User.objects.filter(branch=branch, role='admin', is_active=True).count()
        
        # Guruhlar soni (faol guruhlar)
        groups_count = Group.objects.filter(branch=branch, is_faol=True).count()
        
        # O'quvchilar soni (Faqat faol guruhlarda borlar, kutish zalidagilar qo'shilmaydi)
        students_count = Student.objects.filter(branch=branch, enrollments__is_active=True).distinct().count()

        # 2. Kutilayotgan daromad (Expected Income)
        # Har bir faol guruhning (monthly_price * o'quvchilar soni)
        groups_expected = Group.objects.filter(
            branch=branch,
            is_faol=True
        ).annotate(
            student_count=Count('students', distinct=True),
            group_expected=F('monthly_price') * F('student_count')
        ).aggregate(
            total_expected=Sum('group_expected')
        )
        expected_income = groups_expected['total_expected'] or 0

        # 3. Qabul qilingan daromad (Received Income - Ledger dan)
        received_income = FinanceTransaction.objects.filter(
            branch=branch,
            date__month=month,
            date__year=year,
            transaction_type='income'
        ).aggregate(
            total=Sum('amount')
        )['total'] or 0

        # 4. Har bir guruh bo'yicha batafsil ma'lumot
        groups_detail = []
        groups_qs = Group.objects.filter(
            branch=branch,
            is_faol=True
        ).annotate(
            student_count=Count('students', distinct=True)
        )

        for group in groups_qs:
            # Guruhning kutilayotgan daromadi
            group_expected = group.monthly_price * group.student_count
            
            # Guruhning qabul qilgan daromadi (shu oy uchun)
            group_received = Payment.objects.filter(
                group=group,
                month__month=month,
                month__year=year,
                is_paid=True
            ).aggregate(
                total=Sum('amount')
            )['total'] or 0
            
            # Qarzdorlik
            group_debt = Payment.objects.filter(
                group=group,
                month__month=month,
                month__year=year,
                is_paid=False
            ).aggregate(
                total=Sum('amount')
            )['total'] or 0

            group_refunds = 0
            for student in group.students.all():
                group_refunds += float(student.calculate_refund_amount(year, month))

            # Darslar soni va kunlik narx 
            lessons_count = len(group.get_lesson_dates(year, month))
            daily_price = group.get_daily_price(year, month)

            groups_detail.append({
                "id": group.id,
                "name": group.name,
                "monthly_price": float(group.monthly_price),
                "student_count": group.student_count,
                "lessons_count": lessons_count,
                "daily_price": float(daily_price),
                "expected_income": float(group_expected),
                "received_income": float(group_received),
                "refund_amount": group_refunds,
                "real_income": float(group_received) - group_refunds,
                "debt": float(group_debt),
                "mentor": group.mentor.get_full_name() if group.mentor else "Yo'q"
            })

        # 5. Xarajatlar (Expenses - Ledger dan)
        expenses = FinanceTransaction.objects.filter(
            branch=branch,
            date__month=month,
            date__year=year,
            transaction_type='expense'
        ).aggregate(
            total=Sum('amount')
        )['total'] or 0

        # Refundlar (faqat ko'rsatish uchun, expense ichiga kirgan bo'ladi)
        total_branch_refunds = FinanceTransaction.objects.filter(
            branch=branch,
            category='refund',
            date__year=year,
            date__month=month
        ).aggregate(total=Sum('amount'))['total'] or 0
        if total_branch_refunds == 0:
            total_branch_refunds = sum(g['refund_amount'] for g in groups_detail)

        # 6. Yakuniy ma'lumot
        data = {
            "branch": {
                "id": branch.id,
                "name": branch.name,
                "address": branch.address
            },
            "stats": {
                "mentors": mentors_count,
                "admins": admins_count,
                "groups": groups_count,
                "students": students_count,
                "attendance_today": {
                    "present": Attendance.objects.filter(group__branch=branch, date=today, is_present=True).count(),
                    "absent": Attendance.objects.filter(group__branch=branch, date=today, is_present=False).count(),
                    "total": Attendance.objects.filter(group__branch=branch, date=today).count()
                }
            },
            "finance": {
                "expected_income": float(expected_income),
                "received_income": float(received_income),
                "refunds": float(total_branch_refunds),
                "real_revenue": float(received_income) - float(total_branch_refunds),
                "debt": float(expected_income - received_income),
                "expenses": float(expenses),
                "net_profit": float(received_income) - float(total_branch_refunds) - float(expenses)
            },
            "groups": groups_detail,
            "period": {
                "month": month,
                "year": year
            }
        }

        return Response(data)


from rest_framework.pagination import PageNumberPagination

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 100

class FinanceTransactionViewSet(viewsets.ModelViewSet):
    """
    Kirim-chiqim operatsiyalari tarixi va boshqaruvi.
    Indexlangan va yuqori unumdorlik uchun optimallashtirilgan.
    """
    queryset = FinanceTransaction.objects.select_related('marked_by', 'branch').all()
    serializer_class = FinanceTransactionSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'finance'
    pagination_class = StandardResultsSetPagination

    def perform_create(self, serializer):
        # Admin yoki SuperAdmin yaratayotganda branch'ni avtomatik to'ldirish
        branch = serializer.validated_data.get('branch')
        if not branch and self.request.user.branch:
            branch = self.request.user.branch
            
        serializer.save(
            marked_by=self.request.user,
            branch=branch
        )
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        # ✅ Umumiy statistika (Filtr bo'yicha)
        from django.db.models import Sum
        totals = queryset.aggregate(
            income=Sum('amount', filter=Q(transaction_type='income')),
            expense=Sum('amount', filter=Q(transaction_type='expense'))
        )
        
        income = totals.get('income') or 0
        expense = totals.get('expense') or 0

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            # Custom metadata qo'shamiz
            response.data['stats'] = {
                'income': float(income),
                'expense': float(expense),
                'net': float(income - expense)
            }
            return response

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'results': serializer.data,
            'stats': {
                'income': float(income),
                'expense': float(expense),
                'net': float(income - expense)
            }
        })
    
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # Range-based filtering for date and amount
    filterset_fields = {
        'transaction_type': ['exact'],
        'category': ['exact'],
        'branch': ['exact'],
        'date': ['exact', 'gte', 'lte'],
        'amount': ['gte', 'lte'],
    }
    
    # Indexed text search
    search_fields = ['title', 'description', 'related_id']
    
    ordering_fields = ['date', 'amount', 'created_at']
    ordering = ['-date', '-id']
    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        
        # Adminlar faqat o'z filialini ko'radi (SuperAdmin hammani)
        if user.role == 'admin':
            return qs.filter(branch=user.branch)
        elif user.role != 'super_admin':
            # Faqat admin va super_admin ko'ra oladi
            return FinanceTransaction.objects.none()
            
        return qs

class EmployeeAdvanceViewSet(viewsets.ModelViewSet):
    """Xodimlar avanslarini boshqarish"""
    queryset = EmployeeAdvance.objects.select_related('employee', 'marked_by').all()
    serializer_class = EmployeeAdvanceSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    module_name = 'finance'
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'super_admin':
            return self.queryset
        elif user.role == 'admin':
            return self.queryset.filter(employee__branch=user.branch)
        return self.queryset.filter(employee=user)

    def perform_destroy(self, instance):
        # Ledgerdan ham o'chirish (ixtiyoriy, lekin yaxshi)
        try:
            from .models import FinanceTransaction
            FinanceTransaction.objects.filter(related_id=f"ADV-{instance.id}").delete()
        except:
            pass
        instance.delete()

from rest_framework.pagination import PageNumberPagination

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

class AbsentTodayStudentsView(APIView):
    """
    Bugun kelmagan o'quvchilarni qidirish va paginatsiya bilan qaytaradi.
    GET /finance/statistics/absent-students/<branch_id>/?search=...&page=...
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, branch_id):
        user = request.user
        
        # Ruxsat tekshirish
        if user.role == 'admin':
            has_access = False
            if user.branch_id == branch_id:
                has_access = True
            elif hasattr(user, 'branch_accesses') and user.branch_accesses.filter(branch_id=branch_id).exists():
                has_access = True
            
            if not has_access:
                return Response({"error": "Ruxsat yo'q"}, status=403)
        elif user.role != 'super_admin':
            return Response({"error": "Ruxsat yo'q"}, status=403)

        today = timezone.localdate()
        search = request.query_params.get('search', '').strip()
        export_mode = request.query_params.get('export', 'json')
        
        queryset = Attendance.objects.filter(
            group__branch_id=branch_id, 
            date=today, 
            is_present=False
        ).select_related('student', 'group')

        if search:
            queryset = queryset.filter(
                Q(student__full_name__icontains=search) |
                Q(student__phone__icontains=search) |
                Q(group__name__icontains=search)
            )

        queryset = queryset.order_by('student__full_name')

        if export_mode == 'excel':
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.http import HttpResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bugun Kelmaganlar"

            # Header Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid") # Red for absence
            center_align = Alignment(horizontal='center', vertical='center')

            headers = ["№", "O'quvchi ismi-familiyasi", "Guruh", "Telefon"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20

            # Data
            for row_num, att in enumerate(queryset, 2):
                ws.cell(row=row_num, column=1).value = row_num - 1
                ws.cell(row=row_num, column=2).value = att.student.full_name if att.student else "Noma'lum"
                ws.cell(row=row_num, column=3).value = att.group.name if att.group else "Noma'lum"
                ws.cell(row=row_num, column=4).value = att.student.phone if att.student else ""

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="kelmaganlar_{today}.xlsx"'
            wb.save(response)
            return response

        paginator = StandardResultsSetPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request, view=self)

        results = [
            {
                "id": att.student.id if att.student else None,
                "name": att.student.full_name if att.student else "Noma'lum",
                "phone": att.student.phone if att.student else "",
                "group": att.group.name if att.group else ""
            }
            for att in paginated_queryset
        ]

        return paginator.get_paginated_response(results)
