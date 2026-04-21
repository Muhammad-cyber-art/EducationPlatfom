from rest_framework.viewsets import ModelViewSet
from rest_framework.exceptions import PermissionDenied
from rest_framework.decorators import action
from django.db import transaction,models
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, mixins, status, permissions
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.utils import timezone
from datetime import date ,timedelta
from .models import Attendance , Homework, HomeworkSubmission, MockTest, MockTestResult
from groups.models import Group, Student
from .serializers import (
    HomeworkListSerializer,      # Yangi nom
    HomeworkDetailSerializer,    # Yangi nom
    HomeworkSubmissionUpdateSerializer,
    MockTestListSerializer,
    MockTestDetailSerializer,
    MockTestResultUpdateSerializer,
    AttendanceSerializer,
)
from permissions.permissions import HasModulePermission
from telegram_bot.signals import send_attendance_notification
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from calendar import monthrange

class HomeworkViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'homework'
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['group'] 

    def get_queryset(self):
        user = self.request.user
        group_id = self.request.query_params.get('group_id')
        
        if user.role == 'super_admin':
            queryset = Homework.objects.all()
        elif user.role == 'admin':
            queryset = Homework.objects.filter(group__branch=user.branch)
        else:
            # Mentor access
            group_ids = list(Group.objects.filter(mentor=user).values_list('id', flat=True))
            group_ids += list(Group.objects.filter(additional_mentors__mentor=user).values_list('id', flat=True))
            queryset = Homework.objects.filter(group_id__in=group_ids)

        if group_id:
            queryset = queryset.filter(group_id=group_id)

        if self.action == 'retrieve':
            queryset = queryset.prefetch_related('submissions__student')

        return queryset.select_related('group').distinct()

    def get_serializer_class(self):
        if self.action == 'list':
            return HomeworkListSerializer
        if self.action == 'retrieve':
            return HomeworkDetailSerializer
        return HomeworkListSerializer

    def perform_create(self, serializer):
        user = self.request.user
        group_id = self.request.data.get('group')

        # Permission check for group
        query = models.Q(id=group_id)
        if user.role == 'mentor':
            query &= (models.Q(mentor=user) | models.Q(additional_mentors__mentor=user))
        elif user.role == 'admin':
            query &= models.Q(branch=user.branch)

        group = get_object_or_404(Group, query)

        with transaction.atomic():
            homework = serializer.save(mentor=user, group=group)
            students = group.students.all()
            
            submissions = [
                HomeworkSubmission(
                    homework=homework,
                    student=student,
                    status=HomeworkSubmission.NOT_SUBMITTED
                )
                for student in students
            ]
            HomeworkSubmission.objects.bulk_create(submissions)

    def perform_destroy(self, instance):
        """Uyga vazifani o'chirishdan oldin arxivlash"""
        from archivebase.models import ArchivedHomework
        
        # 1. Statistikani hisoblash
        submissions = instance.submissions.all().select_related('student')
        total = submissions.count()
        full = submissions.filter(status='full').count()
        half = submissions.filter(status='half').count()
        not_sub = submissions.filter(status='not_submitted').count()
        
        submission_rate = (full / total * 100) if total > 0 else 0
        quality_rate = ((full + (half * 0.5)) / total * 100) if total > 0 else 0
        
        stats = {
            "total_students": total,
            "full_submissions": full,
            "half_submissions": half,
            "not_submitted": not_sub,
            "submission_rate": round(submission_rate, 2),
            "quality_rate": round(quality_rate, 2),
            # Har bir o'quvchi ma'lumotlarini ham saqlaymiz
            "students_data": [
                {
                    "name": sub.student.full_name,
                    "status": sub.get_status_display(),
                    "date": sub.submitted_at.strftime('%Y-%m-%d %H:%M') if sub.submitted_at else None
                }
                for sub in submissions
            ]
        }
        
        # 2. Arxivlash
        ArchivedHomework.objects.create(
            original_id=instance.id,
            full_name=instance.title,
            item_type='homework',
            archived_by=self.request.user,
            group_id=instance.group.id,
            group_name=instance.group.name,
            mentor_name=instance.mentor.get_full_name() if instance.mentor else "Noma'lum",
            submission_stats=stats,
            metadata={
                "description": instance.description,
                "created_at": instance.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                "group_id": instance.group.id
            }
        )
        
        # 3. Haqiqiy o'chirish
        instance.delete()

    @action(detail=True, methods=['patch'])
    def update_student_status(self, request, pk=None):
        submission_id = request.data.get('submission_id')
        new_status = request.data.get('status')
        
        try:
            sub = HomeworkSubmission.objects.get(
                    id=submission_id,
                    homework_id=pk
                )
            serializer = HomeworkSubmissionUpdateSerializer(sub, data={'status': new_status}, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({"status": "updated", "new_status": new_status})
        except (HomeworkSubmission.DoesNotExist, ValueError, TypeError):
            return Response({"error": "Topshiriq topilmadi"}, status=404)

# =================================== ATTENDS ===============================

class AttendanceViewSet(viewsets.ModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'homework'
    http_method_names = ['get', 'post', 'patch', 'head', 'options']

    def get_queryset(self):
       group_id = self.request.query_params.get('group_id')
       date_str = self.request.query_params.get('date', str(timezone.localdate()))
       today = timezone.localdate()
       
       if not group_id:
           return Attendance.objects.none()
       
       try:
           requested_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
       except (ValueError, TypeError):
           requested_date = today

       group = get_object_or_404(Group, id=group_id)
       
       # Dars kuni ekanligini tekshiramiz
       is_lesson_day = group.is_lesson_day(requested_date)
       
       if not is_lesson_day:
           return Attendance.objects.none()
       
       queryset = Attendance.objects.filter(group=group, date=requested_date)

       active_students = group.students.filter(joined_at__date__lte=requested_date)

       if is_lesson_day and queryset.count() < active_students.count():
           existing_student_ids = queryset.values_list('student_id', flat=True)
           new_records = [
               Attendance(student=student, group=group, date=requested_date, is_present=True)
               for student in active_students if student.id not in existing_student_ids
           ]
           if new_records:
               Attendance.objects.bulk_create(new_records)
           queryset = Attendance.objects.filter(group=group, date=requested_date)

       return queryset.filter(student__joined_at__date__lte=requested_date).select_related('student').order_by('student__full_name')

    @action(detail=False, methods=['get'])
    def weekly_report(self, request):
         group_id = request.query_params.get('group_id')
         group = get_object_or_404(Group, id=group_id)

         today = timezone.localdate()
         start_of_week = today - timedelta(days=today.weekday()) 
         date_list = [(start_of_week + timedelta(days=i)) for i in range(7)]

         students = group.students.all().order_by('full_name')
         attendances = Attendance.objects.filter(
             group=group, 
             date__range=[date_list[0], date_list[-1]]
         )

         att_data = {}
         for att in attendances:
             if att.student_id not in att_data:
                 att_data[att.student_id] = {}
             att_data[att.student_id][str(att.date)] = att.is_present

         report = []
         for student in students:
             student_history = []
             joined_date = student.joined_at.date() if student.joined_at else None

             for d in date_list:
                 date_str = str(d)
                 if joined_date and d < joined_date:
                     status = None
                 elif not group.is_lesson_day(d):
                     status = None
                 else:
                     status = att_data.get(student.id, {}).get(date_str, True)

                 student_history.append({
                     "date": date_str, 
                     "day_name": d.strftime('%A'),
                     "is_present": status
                 })

             report.append({
                 "student_id": student.id,
                 "student_name": student.full_name,
                 "history": student_history
             })

         return Response({
             "week_start": str(date_list[0]),
             "week_end": str(date_list[-1]),
             "data": report
         })
    
    def create(self, request, *args, **kwargs):
        attendance_id = request.data.get('id')
        is_present = request.data.get('is_present')
        student_id = request.data.get('student_id')
        date_str = request.data.get('date')

        # 1. Update mavjud yozuvni
        if attendance_id is not None:
            instance = get_object_or_404(Attendance, id=attendance_id)
            instance.is_present = is_present
            instance.marked_by = request.user
            instance.save()
            serializer = self.get_serializer(instance)
            return Response(serializer.data, status=status.HTTP_200_OK)

        # 2. Yangi yozuv yaratish
        if not (student_id and date_str):
             return Response({"detail": "Student ID va kun (date) majburiy"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            requested_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"detail": "Sana formati noto'g'ri (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)

        group_id = request.query_params.get('group_id')
        group = get_object_or_404(Group, id=group_id)
        
        if not group.is_lesson_day(requested_date):
            return Response({"detail": "Tanlangan sana guruh uchun dars kuni emas"}, status=status.HTTP_400_BAD_REQUEST)
            
        student = get_object_or_404(group.students, id=student_id)

        # Duplikat tekshirish
        attendance, created = Attendance.objects.get_or_create(
            student=student, 
            group=group, 
            date=requested_date,
            defaults={
                'is_present': is_present,
                'marked_by': request.user
            }
        )
        if not created:
            attendance.is_present = is_present
            attendance.marked_by = request.user
            attendance.save()

        serializer = self.get_serializer(attendance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='confirm')
    def confirm_attendance(self, request):
        """
        Admin davomatlarni belgilab «Tayyor» bosganda chaqiriladi.
        Barcha davomatlarni optimal tarzda (bulk) saqlaydi va 
        xabarnomalarni fon vazifasi (Celery) orqali yuboradi.
        """
        group_id = request.data.get('group_id')
        date_str = request.data.get('date')
        attendances_payload = request.data.get('attendances', [])

        if not group_id or not date_str:
            return Response({"detail": "group_id va date majburiy"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            requested_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"detail": "Sana formati noto'g'ri (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)

        group = get_object_or_404(Group, id=group_id)
        
        if not group.is_lesson_day(requested_date):
            return Response({"detail": "Ushbu sana guruh uchun dars kuni emas"}, status=status.HTTP_400_BAD_REQUEST)
        
        # 1. Mavjud davomatlarni bir marta olib kelamiz
        existing_attendances = Attendance.objects.filter(group=group, date=requested_date)
        attendance_map = {att.student_id: att for att in existing_attendances}
        
        to_create = []
        to_update = []
        updated_ids = []

        with transaction.atomic():
            for item in attendances_payload:
                student_id = item.get('student_id')
                is_present = item.get('is_present', True)
                if student_id is None: continue

                if student_id in attendance_map:
                    attendance = attendance_map[student_id]
                    if attendance.is_present != is_present or attendance.marked_by != request.user:
                        attendance.is_present = is_present
                        attendance.marked_by = request.user
                        to_update.append(attendance)
                    updated_ids.append(attendance.id)
                else:
                    # Agar bazada hali yo'q bo'lsa (masalan yangi student qo'shilgan)
                    to_create.append(Attendance(
                        student_id=student_id,
                        group=group,
                        date=requested_date,
                        is_present=is_present,
                        marked_by=request.user
                    ))

            # 2. Bulk operatsiyalar
            if to_update:
                Attendance.objects.bulk_update(to_update, ['is_present', 'marked_by'])
            if to_create:
                created_objs = Attendance.objects.bulk_create(to_create)
                updated_ids.extend([obj.id for obj in created_objs])

        # 3. Xabarnomalarni fon vazifasiga berish
        try:
            from telegram_bot.tasks import send_attendance_notifications_task
            send_attendance_notifications_task.delay(updated_ids)
        except Exception:
            # Celery/Redis ishlamayotgan bo'lsa, xabarnomalarni bitta alohida thread ichida navbat bilan yuboramiz
            import threading
            from telegram_bot.signals import send_attendance_notification
            def fallback_notifications():
                atts = Attendance.objects.filter(id__in=updated_ids)
                for a in atts: 
                    try:
                        send_attendance_notification(a, async_send=False)
                        import time
                        time.sleep(0.05)
                    except: pass
            threading.Thread(target=fallback_notifications).start()

        # Natijani qaytaramiz
        queryset = Attendance.objects.filter(group=group, date=requested_date).select_related('student').order_by('student__full_name')
        serializer = AttendanceSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='monthly-report')
    def monthly_report(self, request):
        """
        Guruh uchun oylik davomat hisoboti.
        Parametrlar: group_id (majburiy), month, year, export (json/excel)
        """
        group_id = request.query_params.get('group_id')
        if not group_id:
            return Response({"detail": "group_id majburiy"}, status=status.HTTP_400_BAD_REQUEST)
            
        group = get_object_or_404(Group, id=group_id)
        today = timezone.localdate()
        
        # Oyni aniqlaymiz
        try:
            month = int(request.query_params.get('month', today.month))
            year = int(request.query_params.get('year', today.year))
        except (ValueError, TypeError):
            month, year = today.month, today.year
            
        first_day = date(year, month, 1)
        _, last_day_num = monthrange(year, month)
        last_day = date(year, month, last_day_num)
        
        # Agar joriy oy bo'lsa, faqat bugungacha bo'lgan kunlarni ko'rsatamiz
        if year == today.year and month == today.month:
            report_end_day = today
        else:
            report_end_day = last_day
            
        date_list = [first_day + timedelta(days=i) for i in range((report_end_day - first_day).days + 1)]
        
        students = group.students.all().order_by('full_name')
        attendances = Attendance.objects.filter(
            group=group,
            date__range=[first_day, report_end_day]
        )
        
        att_data = {}
        for att in attendances:
            if att.student_id not in att_data:
                att_data[att.student_id] = {}
            att_data[att.student_id][str(att.date)] = att.is_present
            
        export_mode = request.query_params.get('export', 'json')
        
        if export_mode == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Davomat_{group.name[:20]}"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Header row
            headers = ["№", "O'quvchi ismi-familiyasi", "Telefon", "Ota-ona telefon"] + [d.day for d in date_list]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header

            # Column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            for col_num in range(5, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 4
                
            # Header Row Styling
            header_row = ws[2]
            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Data rows
            for row_num, student in enumerate(students, 3):
                ws.cell(row=row_num, column=1).value = row_num - 2
                ws.cell(row=row_num, column=2).value = student.full_name
                ws.cell(row=row_num, column=3).value = student.phone or "-"
                ws.cell(row=row_num, column=4).value = student.parent_phone or "-"
                
                joined_at = student.joined_at.date() if student.joined_at else None
                
                for col_idx, d in enumerate(date_list, 5):
                    cell = ws.cell(row=row_num, column=col_idx)
                    
                    if joined_at and d < joined_at:
                        cell.value = "-"
                    elif not group.is_lesson_day(d):
                        cell.value = "X" 
                        cell.font = Font(color="808080")
                    else:
                        is_present = att_data.get(student.id, {}).get(str(d))
                        if is_present is True:
                            cell.value = "+"
                            cell.font = Font(color="008000", bold=True)
                        elif is_present is False:
                            cell.value = "K" 
                            cell.font = Font(color="FF0000", bold=True)
                        else:
                            cell.value = "?" 
                    cell.alignment = center_align

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="davomat_{group.name}_{year}_{month}.xlsx"'
            wb.save(response)
            return response

        # JSON Response
        report_data = []
        for student in students:
            history = []
            joined_at = student.joined_at.date() if student.joined_at else None
            for d in date_list:
                date_str = str(d)
                if joined_at and d < joined_at:
                    status = "not_joined"
                elif not group.is_lesson_day(d):
                    status = "no_lesson"
                else:
                    is_p = att_data.get(student.id, {}).get(date_str)
                    if is_p is True: status = "present"
                    elif is_p is False: status = "absent"
                    else: status = "none"
                
                history.append({"date": date_str, "status": status})
            
            report_data.append({
                "student_id": student.id,
                "student_name": student.full_name,
                "attendance": history
            })

        return Response({
            "group_name": group.name,
            "period": f"{year}-{month:02d}",
            "data": report_data
        })

class MockTestViewSet(ModelViewSet):
    permission_classes = [IsAuthenticated, HasModulePermission]
    module_name = 'homework'
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['group']

    def get_queryset(self):
        user = self.request.user
        group_id = self.request.query_params.get('group_id')
        
        if user.role == 'super_admin':
            queryset = MockTest.objects.all()
        elif user.role == 'admin':
            queryset = MockTest.objects.filter(group__branch=user.branch)
        else:
            # Mentor access
            group_ids = list(Group.objects.filter(mentor=user).values_list('id', flat=True))
            group_ids += list(Group.objects.filter(additional_mentors__mentor=user).values_list('id', flat=True))
            queryset = MockTest.objects.filter(group_id__in=group_ids)

        if group_id:
            queryset = queryset.filter(group_id=group_id)

        if self.action == 'retrieve':
            queryset = queryset.prefetch_related('results__student')

        return queryset.select_related('group').distinct()

    def get_serializer_class(self):
        if self.action == 'list':
            return MockTestListSerializer
        if self.action == 'retrieve':
            return MockTestDetailSerializer
        return MockTestListSerializer

    def perform_create(self, serializer):
        user = self.request.user
        group_id = self.request.data.get('group')

        # Permission check for group
        query = models.Q(id=group_id)
        if user.role == 'mentor':
            query &= (models.Q(mentor=user) | models.Q(additional_mentors__mentor=user))
        elif user.role == 'admin':
            query &= models.Q(branch=user.branch)

        group = get_object_or_404(Group, query)

        with transaction.atomic():
            mock_test = serializer.save(group=group)
            students = group.students.all()
            
            results = [
                MockTestResult(
                    test=mock_test,
                    student=student,
                    score=''
                )
                for student in students
            ]
            MockTestResult.objects.bulk_create(results)

    def perform_destroy(self, instance):
        """Mock testni o'chirishdan oldin arxivlash"""
        from archivebase.models import ArchivedHomework
        
        # 1. Statistikani hisoblash
        results = instance.results.all().select_related('student')
        total = results.count()
        scored = results.exclude(score='').exclude(score__isnull=True).count()
        
        stats = {
            "total_students": total,
            "participated": scored,
            "not_participated": total - scored,
            "participation_rate": round((scored / total * 100), 2) if total > 0 else 0,
            "students_data": [
                {
                    "name": res.student.full_name,
                    "score": res.score or "N/A",
                    "status": "Qatnashgan" if res.score else "Qatnashmagan"
                }
                for res in results
            ]
        }
        
        # 2. Arxivlash
        ArchivedHomework.objects.create(
            original_id=instance.id,
            full_name=instance.subject, # Mock testda title o'rniga subject
            item_type='mock_test',
            archived_by=self.request.user,
            group_id=instance.group.id,
            group_name=instance.group.name,
            mentor_name=instance.group.mentor.get_full_name() if instance.group.mentor else "Noma'lum",
            submission_stats=stats,
            metadata={
                "test_type": instance.type,
                "test_date": instance.date.strftime('%Y-%m-%d'),
                "created_at": instance.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                "group_id": instance.group.id
            }
        )
        
        # 3. Haqiqiy o'chirish
        instance.delete()

    @action(detail=True, methods=['patch'])
    def update_student_score(self, request, pk=None):
        result_id = request.data.get('result_id')
        score = request.data.get('score')
        
        try:
            result = MockTestResult.objects.get(
                id=result_id,
                test_id=pk
            )
            serializer = MockTestResultUpdateSerializer(result, data={'score': score}, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response({"status": "updated", "new_score": score})
        except (MockTestResult.DoesNotExist, ValueError, TypeError):
            return Response({"error": "Natija topilmadi"}, status=404)