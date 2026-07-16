from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated,AllowAny
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.http import HttpResponse

import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import timedelta

from homework_attends.models import Attendance
from groups.models import Group
from homework_attends.serializers import AttendanceSerializer, AttendanceEditSerializer
from homework_attends.services import (
    get_or_create_attendance_records, 
    generate_weekly_attendance_report, 
    bulk_confirm_attendance, 
    get_monthly_attendance_data,
    edit_past_attendance,
    get_last_3_lesson_dates
)
from permissions.permissions import HasModulePermission
from reports.models import ReportDownloadTrack
from datetime import date as date_type

logger = logging.getLogger(__name__)

class AttendanceViewSet(viewsets.ModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [AllowAny]
    pagination_class = None
    module_name = 'homework'
    http_method_names = ['get', 'post', 'patch', 'delete', 'head', 'options']

    def get_permissions(self):
        """Monthly report uchun autentifikatsiya majburiy (ReportDownloadTrack user talab qiladi)"""
        if self.action == 'monthly_report':
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
       group_id = self.request.query_params.get('group_id')
       date_str = self.request.query_params.get('date', str(timezone.localdate()))
       
       if not group_id:
           return Attendance.objects.none()
       
       try:
           requested_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
       except (ValueError, TypeError):
           requested_date = timezone.localdate()

       group = get_object_or_404(Group, id=group_id)
       # O'tgan oylar uchun: faqat mavjud yozuvlarni ko'rsatish (yangi yaratmaslik)
       today = timezone.localdate()
       is_past_month = requested_date.year < today.year or (requested_date.year == today.year and requested_date.month < today.month)
       return get_or_create_attendance_records(group, requested_date, view_only=is_past_month)

    @action(detail=False, methods=['get'])
    def weekly_report(self, request):
         group_id = request.query_params.get('group_id')
         group = get_object_or_404(Group, id=group_id)
         
         report = generate_weekly_attendance_report(group)
         return Response(report)
    
    def create(self, request, *args, **kwargs):
        attendance_id = request.data.get('id')
        is_present = request.data.get('is_present')
        student_id = request.data.get('student_id')
        date_str = request.data.get('date')

        # 1. Update existing record
        if attendance_id is not None:
            instance = get_object_or_404(Attendance, id=attendance_id)
            today = timezone.localdate()
            if instance.date != today:
                last_3_dates = get_last_3_lesson_dates(instance.group)
                
                # Grace period: Tungi 4 gacha kechagi kunni o'zgartirishga ruxsat
                is_early_morning = timezone.localtime().hour < 4
                is_yesterday = instance.date == (today - timedelta(days=1))
                if instance.date not in last_3_dates and not (is_early_morning and is_yesterday):
                    return Response({"detail": f"Faqat bugungi yoki oxirgi 3 ta o'tilgan dars davomatini tahrirlashingiz mumkin."}, status=status.HTTP_400_BAD_REQUEST)
            
            instance.is_present = is_present
            instance.marked_by = request.user
            instance.save()
            serializer = self.get_serializer(instance)
            return Response(serializer.data, status=status.HTTP_200_OK)

        # 2. Create new record
        if not (student_id and date_str):
             return Response({"detail": "Student ID va kun (date) majburiy"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            requested_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"detail": "Sana formati noto'g'ri (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)

        group = get_object_or_404(Group, id=group_id)
        
        # Ruxsat etilgan sanalar: Bugun, yoki oxirgi 3 ta dars
        today = timezone.localdate()
        if requested_date != today:
            last_3_dates = get_last_3_lesson_dates(group)
            
            is_early_morning = timezone.localtime().hour < 4
            is_yesterday = requested_date == (today - timedelta(days=1))
            if requested_date not in last_3_dates and not (is_early_morning and is_yesterday):
                return Response({"detail": f"Faqat bugungi yoki oxirgi 3 ta o'tilgan dars davomatini tahrirlashingiz mumkin."}, status=status.HTTP_400_BAD_REQUEST)

        student = get_object_or_404(group.students, id=student_id)

        # Duplicate check
        attendance, created = Attendance.objects.get_or_create(
            student=student, 
            group=group, 
            date=requested_date,
            defaults={
                'is_present': is_present,
                'marked_by': request.user
            }
        )
        if not created:
            attendance.is_present = is_present
            attendance.marked_by = request.user
            attendance.save()

        serializer = self.get_serializer(attendance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='confirm')
    def confirm_attendance(self, request):
        group_id = request.data.get('group_id')
        date_str = request.data.get('date')
        attendances_payload = request.data.get('attendances', [])

        if not group_id or not date_str:
            return Response({"detail": "group_id va date majburiy"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            requested_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"detail": "Sana formati noto'g'ri (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST)

        group = get_object_or_404(Group, id=group_id)
        
        # Ruxsat etilgan sanalar: Bugun, yoki oxirgi 3 ta dars
        today = timezone.localdate()
        if requested_date != today:
            last_3_dates = get_last_3_lesson_dates(group)
            
            is_early_morning = timezone.localtime().hour < 4
            is_yesterday = requested_date == (today - timedelta(days=1))
            if requested_date not in last_3_dates and not (is_early_morning and is_yesterday):
                return Response({"detail": f"Faqat bugungi yoki oxirgi 3 ta o'tilgan dars davomatini tahrirlashingiz mumkin."}, status=status.HTTP_400_BAD_REQUEST)


        
        queryset = bulk_confirm_attendance(group, requested_date, attendances_payload, request.user)
        serializer = AttendanceSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def _user_can_access_group(self, user, group):
        """Foydalanuvchi guruhga kirish huquqini tekshirish"""
        if user.role == 'super_admin':
            return True
        if user.role == 'admin':
            allowed = [user.branch_id] if user.branch_id else []
            if hasattr(user, 'branch_accesses'):
                allowed.extend(user.branch_accesses.values_list('branch_id', flat=True))
            return group.branch_id in allowed
        if user.role == 'mentor':
            return group.mentor == user or group.additional_mentors.filter(mentor=user).exists()
        return False

    @action(detail=False, methods=['get'], url_path='monthly-report')
    def monthly_report(self, request):
        group_id = request.query_params.get('group_id')
        if not group_id:
            return Response({"detail": "group_id majburiy"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            group = get_object_or_404(Group, id=group_id)
            
            # Foydalanuvchining guruhga kirish huquqini tekshirish
            if not self._user_can_access_group(request.user, group):
                logger.warning(
                    "User %s (role=%s) has no access to group %s (branch=%s)",
                    request.user, request.user.role, group.id, group.branch_id
                )
                return Response(
                    {"detail": "Sizda ushbu guruh davomatini ko'rish huquqi yo'q"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            today = timezone.localdate()
            
            try:
                month = int(request.query_params.get('month', today.month))
                year = int(request.query_params.get('year', today.year))
                # Oy va yil qiymatlarini tekshirish
                date_type(year, month, 1)
            except (ValueError, TypeError):
                month, year = today.month, today.year
                
            date_list, students, att_data = get_monthly_attendance_data(group, month, year)
        except Exception as e:
            logger.exception("Monthly report data olishda xatolik: group=%s, month=%s-%s", group_id, year, month)
            return Response(
                {"detail": f"Hisobot ma'lumotlarini olishda xatolik: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        export_mode = request.query_params.get('export', 'json')
        
        if export_mode == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # openpyxl sheet title has strict rules: max 31 chars, no forbidden chars: \ / ? * : [ ]
            safe_sheet_name = f"Davomat_{group.name[:20]}"
            for char in r"\/?:*[]":
                safe_sheet_name = safe_sheet_name.replace(char, "")
            safe_sheet_name = safe_sheet_name.strip()[:30]
            if not safe_sheet_name:
                safe_sheet_name = "Davomat"
            ws.title = safe_sheet_name
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            
            headers = ["№", "O'quvchi ismi-familiyasi", "Telefon", "Ota-ona telefon"] + [d.day for d in date_list]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            for col_num in range(5, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 4
                
            header_row = ws[2]
            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            for row_num, student in enumerate(students, 3):
                ws.cell(row=row_num, column=1).value = row_num - 2
                ws.cell(row=row_num, column=2).value = student.full_name
                ws.cell(row=row_num, column=3).value = student.phone or "-"
                ws.cell(row=row_num, column=4).value = student.parent_phone or "-"
                
                joined_at = student.joined_at.date() if student.joined_at else None
                
                for col_idx, d in enumerate(date_list, 5):
                    cell = ws.cell(row=row_num, column=col_idx)
                    
                    att_record = att_data.get(student.id, {}).get(str(d))
                    
                    if att_record:
                        if not att_record.get('is_confirmed'):
                            cell.value = "!" # Davomat olinmagan (tasdiqlanmagan)
                            cell.font = Font(color="FF0000", bold=True)
                        elif att_record.get('is_present') is True:
                            cell.value = "+"
                            cell.font = Font(color="008000", bold=True)
                        else:
                            cell.value = "K" 
                            cell.font = Font(color="FF0000", bold=True)
                    elif joined_at and d < joined_at:
                        cell.value = "-"
                    else:
                        # Record yo'q bo'lsa va dars kuni bo'lsa (date_list dars kunlaridan iborat)
                        cell.value = "!"
                        cell.font = Font(color="FF0000", bold=True)

                    cell.alignment = center_align

            # Yuklab olish tarixini saqlash (xatolik bo'lsa ham report yuklanishi kerak)
            try:
                report_date = date_type(year, month, 1)
                ReportDownloadTrack.objects.get_or_create(
                    user=request.user,
                    report_type='attendance_monthly',
                    report_date=report_date
                )
            except Exception as e:
                logger.warning(
                    "ReportDownloadTrack saqlashda xatolik: user=%s, group=%s, month=%s-%s — %s",
                    request.user, group.id, year, month, e
                )

            from urllib.parse import quote
            # Format filename safely to avoid UnicodeEncodeError in Content-Disposition header
            raw_filename = f"davomat_{group.name}_{year}_{month}.xlsx"
            # Remove characters that can break the HTTP header or represent directory traversal
            safe_filename = raw_filename.replace('\n', '').replace('\r', '').replace('"', '_').replace('/', '_').replace('\\', '_')
            encoded_filename = quote(safe_filename)
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"; filename*=utf-8\'\'{encoded_filename}'
            try:
                wb.save(response)
            except Exception as e:
                logger.exception("Excel faylni yaratishda xatolik: group=%s, month=%s-%s", group.id, year, month)
                return Response(
                    {"detail": f"Excel faylni yaratishda xatolik: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            return response

        # JSON Response
        report_data = []
        for student in students:
            history = []
            joined_at = student.joined_at.date() if student.joined_at else None
            for d in date_list:
                date_str = str(d)
                att_record = att_data.get(student.id, {}).get(date_str)
                
                # BUG #6 FIX: 'att_status' nomini ishlatamiz — 'status' DRF moduli bilan to'qnashmaslik uchun
                if joined_at and d < joined_at:
                    att_status = "not_joined"
                elif att_record:
                    if not att_record.get('is_confirmed'):
                        att_status = "not_taken"
                    elif att_record.get('is_present'):
                        att_status = "present"
                    else:
                        att_status = "absent"
                else:
                    # Dars kuni, lekin rekord yo'q
                    if d < timezone.localdate():
                        att_status = "not_taken"
                    else:
                        att_status = "none"
                
                history.append({"date": date_str, "status": att_status})
            
            report_data.append({
                "student_id": student.id,
                "student_name": student.full_name,
                "attendance": history
            })

        return Response({
            "group_name": group.name,
            "period": f"{year}-{month:02d}",
            "data": report_data
        })

    @action(detail=True, methods=['patch'], url_path='edit-past', permission_classes=[IsAuthenticated])
    def edit_past(self, request, pk=None):
        """
        Oxirgi 3 ta o'tilgan darsdan birortasining davomatini tahrirlash.
        """
        attendance = self.get_object()
        serializer = AttendanceEditSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        # Admin huquqini tekshirish
        if request.user.role not in ['admin', 'super_admin']:
            return Response(
                {"detail": "Faqat administratorlar o'tgan davomatni tahrirlay oladi."},
                status=status.HTTP_403_FORBIDDEN
            )
            
        try:
            from rest_framework.exceptions import ValidationError
            updated_attendance = edit_past_attendance(
                attendance_id=attendance.id,
                admin_user=request.user,
                new_status=serializer.validated_data['new_status'],
                reason=serializer.validated_data['reason']
            )
            return Response(
                AttendanceSerializer(updated_attendance).data, 
                status=status.HTTP_200_OK
            )
        except ValidationError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception("Davomat tahrirlashda xatolik: %s", str(e))
            return Response(
                {"detail": "Tizim xatoligi yuz berdi. Iltimos adminga murojaat qiling."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
