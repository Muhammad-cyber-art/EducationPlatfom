"""
ReportGenerator Service - SOLID Principles Compliant

This module provides a comprehensive reporting system for the CRM-integrated Telegram bot.
It handles daily, monthly attendance, and financial reports with proper error handling and logging.

Architecture:
- Single Responsibility: Each method handles one specific report type
- Open/Closed: Extensible for new report types without modifying existing code
- Liskov Substitution: Report generators can be swapped
- Interface Segregation: Focused interfaces for different report types
- Dependency Inversion: Depends on abstractions (Django models) not concrete implementations
"""

import logging
import io
import zipfile
from datetime import datetime, date
from typing import Optional, List, Dict, Tuple
from django.utils import timezone
from django.db.models import Sum, F, Q
from django.conf import settings

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from homework_attends.models import Attendance, Homework, HomeworkSubmission
from groups.models import Group, Student
from finance.models import Payment, EmployeePayment, FinanceTransaction
from branches.models import Branch

logger = logging.getLogger(__name__)


class ReportGenerationError(Exception):
    """Custom exception for report generation errors"""
    pass


class ReportGenerator:
    """
    Main report generator service class.
    Handles all report generation with proper error handling and logging.
    """
    
    # Excel styling constants
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def __init__(self):
        self.logger = logger
    
    def _apply_header_style(self, row, ws):
        """Apply consistent header styling to a row"""
        for cell in row:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.BORDER
    
    def _apply_row_style(self, row, ws):
        """Apply consistent row styling"""
        for cell in row:
            cell.border = self.BORDER
            cell.alignment = Alignment(vertical="center", horizontal="left")
    
    def _auto_adjust_columns(self, ws, min_width=12, max_width=50):
        """Auto-adjust column widths based on content"""
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value:
                        val_len = len(str(cell.value))
                        if val_len > max_length:
                            max_length = val_len
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 3, min_width), max_width)
    
    def generate_daily_branch_report(self, branch: Branch, target_date: Optional[date] = None) -> io.BytesIO:
        """
        Generate daily Excel report for a specific branch.
        
        Args:
            branch: Branch object to generate report for
            target_date: Date for the report (defaults to today)
            
        Returns:
            BytesIO object containing the Excel file
            
        Raises:
            ReportGenerationError: If report generation fails
        """
        try:
            if target_date is None:
                target_date = timezone.now().date()
            
            self.logger.info(f"Generating daily report for branch {branch.name} on {target_date}")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Daily Report {target_date.strftime('%Y.%m.%d')}"
            
            # Headers
            headers = [
                "№", "Guruh", "O'quvchi F.I.SH", 
                "Davomat", "Uy vazifasi (Mavzu)", "Vazifa holati"
            ]
            ws.append(headers)
            self._apply_header_style(ws[1], ws)
            
            # Query attendances for this branch
            attendances = (
                Attendance.objects
                .filter(
                    date=target_date,
                    group__branch=branch,
                    group__isnull=False,
                    student__isnull=False,
                    group__is_archived=False,
                )
                .select_related('student', 'group')
                .prefetch_related('group__homeworks')
                .order_by('group__name', 'student__full_name')
            )
            
            if not attendances.exists():
                self.logger.warning(f"No attendance data found for branch {branch.name} on {target_date}")
                ws.append(["Ma'lumot topilmadi"])
            else:
                # Cache homework by group
                homework_by_group = {}
                
                for index, att in enumerate(attendances, start=1):
                    student = att.student
                    group = att.group
                    
                    # Get homework for this group
                    if group.id not in homework_by_group:
                        hw = Homework.objects.filter(
                            group=group,
                            created_at__date=target_date
                        ).first()
                        homework_by_group[group.id] = hw
                    
                    homework = homework_by_group[group.id]
                    
                    hw_title = "Vazifa berilmagan"
                    hw_status = "-"
                    
                    if homework:
                        hw_title = homework.title
                        submission = HomeworkSubmission.objects.filter(
                            homework=homework,
                            student=student
                        ).first()
                        
                        if submission:
                            hw_status = submission.get_status_display()
                        else:
                            hw_status = "Topshirmagan ❌"
                    
                    row_data = [
                        index,
                        group.name,
                        student.full_name,
                        "✅ Keldi" if att.is_present else "❌ Kelmadi",
                        hw_title,
                        hw_status
                    ]
                    ws.append(row_data)
                    self._apply_row_style(ws[ws.max_row], ws)
            
            # Auto-adjust columns
            dims = {'A': 5, 'B': 20, 'C': 35, 'D': 15, 'E': 30, 'F': 20}
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            self.logger.info(f"Successfully generated daily report for branch {branch.name}")
            return output
            
        except Exception as e:
            self.logger.error(f"Error generating daily report for branch {branch.name}: {str(e)}")
            raise ReportGenerationError(f"Failed to generate daily report: {str(e)}")
    
    def generate_monthly_attendance_zip(self, branch: Branch, year: int, month: int) -> io.BytesIO:
        """
        Generate ZIP archive containing daily attendance Excel reports for a month.
        
        Args:
            branch: Branch object to generate reports for
            year: Year for the reports
            month: Month for the reports
            
        Returns:
            BytesIO object containing the ZIP file
            
        Raises:
            ReportGenerationError: If ZIP generation fails
        """
        try:
            self.logger.info(f"Generating monthly attendance ZIP for branch {branch.name} - {year}/{month}")
            
            # Get all days in the month
            if month == 12:
                next_month = date(year + 1, 1, 1)
            else:
                next_month = date(year, month + 1, 1)
            
            current_month = date(year, month, 1)
            days_in_month = (next_month - current_month).days
            
            # Create ZIP in memory
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for day in range(1, days_in_month + 1):
                    target_date = date(year, month, day)
                    
                    try:
                        # Generate daily report for this day
                        excel_buffer = self.generate_daily_branch_report(branch, target_date)
                        
                        # Add to ZIP
                        filename = f"attendance_{target_date.strftime('%Y-%m-%d')}.xlsx"
                        zip_file.writestr(filename, excel_buffer.getvalue())
                        
                        self.logger.debug(f"Added {filename} to ZIP")
                        
                    except Exception as e:
                        self.logger.warning(f"Failed to generate report for {target_date}: {str(e)}")
                        # Continue with other days even if one fails
                        continue
            
            zip_buffer.seek(0)
            
            self.logger.info(f"Successfully generated monthly attendance ZIP for branch {branch.name}")
            return zip_buffer
            
        except Exception as e:
            self.logger.error(f"Error generating monthly attendance ZIP for branch {branch.name}: {str(e)}")
            raise ReportGenerationError(f"Failed to generate monthly attendance ZIP: {str/e}")
    
    def generate_monthly_financial_report(self, year: int, month: int) -> io.BytesIO:
        """
        Generate comprehensive monthly financial Excel report for super_admin.
        
        Args:
            year: Year for the report
            month: Month for the report
            
        Returns:
            BytesIO object containing the Excel file
            
        Raises:
            ReportGenerationError: If report generation fails
        """
        try:
            self.logger.info(f"Generating monthly financial report for {year}/{month}")
            
            wb = openpyxl.Workbook()
            
            # Sheet 1: Student Payments
            ws1 = wb.active
            ws1.title = "O'quvchilar To'lovlari"
            
            headers1 = [
                "№", "Filial", "O'quvchi F.I.SH", "Guruh", "To'lov Summasi",
                "To'lov Oyi", "Tasdiqlagan Admin", "To'langan Vaqt", "Tranzaksiya ID"
            ]
            ws1.append(headers1)
            self._apply_header_style(ws1[1], ws1)
            
            student_payments = (
                Payment.objects
                .filter(
                    month__year=year,
                    month__month=month,
                    is_paid=True,
                    group__isnull=False,
                    group__branch__isnull=False,
                    student__isnull=False,
                )
                .select_related('student', 'group', 'group__branch', 'marked_by')
                .order_by('group__branch', 'student__full_name')
            )
            
            for i, sp in enumerate(student_payments, 1):
                ws1.append([
                    i,
                    sp.group.branch.name if sp.group.branch else "-",
                    sp.student.full_name,
                    sp.group.name,
                    float(sp.amount) if sp.amount else 0,
                    sp.month.strftime("%Y-%m") if sp.month else "-",
                    sp.marked_by.get_full_name() or sp.marked_by.username if sp.marked_by else "Onlayn",
                    sp.paid_at.strftime("%Y-%m-%d %H:%M") if sp.paid_at else "-",
                    sp.transaction_id or "-"
                ])
                self._apply_row_style(ws1[ws1.max_row], ws1)
            
            # Sheet 2: Employee Payments
            ws2 = wb.create_sheet("Xodimlar Maoshlari")
            headers2 = [
                "№", "Filial", "Xodim F.I.SH", "Roli", "Asosiy Maosh",
                "Bonus", "Jarimalar", "JAMI TO'LANDI", "Tasdiqladi", "Sana"
            ]
            ws2.append(headers2)
            self._apply_header_style(ws2[1], ws2)
            
            employee_payments = (
                EmployeePayment.objects
                .filter(month__year=year, month__month=month, is_paid=True)
                .select_related('employee', 'employee__branch', 'marked_by')
                .order_by('employee__branch')
            )
            
            for i, ep in enumerate(employee_payments, 1):
                total_paid = float(ep.salary_base + ep.bonus - ep.deductions)
                ws2.append([
                    i,
                    ep.employee.branch.name if ep.employee.branch else "-",
                    ep.employee.get_full_name() or ep.employee.username,
                    ep.employee.get_role_display(),
                    float(ep.salary_base) if ep.salary_base else 0,
                    float(ep.bonus) if ep.bonus else 0,
                    float(ep.deductions) if ep.deductions else 0,
                    total_paid,
                    ep.marked_by.get_full_name() or ep.marked_by.username if ep.marked_by else "Tizim",
                    ep.paid_at.strftime("%Y-%m-%d") if ep.paid_at else "-"
                ])
                self._apply_row_style(ws2[ws1.max_row], ws2)
            
            # Sheet 3: Other Transactions
            ws3 = wb.create_sheet("Refund va Qo'shimcha Amallar")
            headers3 = [
                "№", "Filial", "Tur", "Turkum", "Sarlavha", "Summa", "Mas'ul", "Sana", "Tavsif"
            ]
            ws3.append(headers3)
            self._apply_header_style(ws3[1], ws3)
            
            other_transactions = (
                FinanceTransaction.objects
                .filter(date__year=year, date__month=month)
                .exclude(category__in=['student_fee', 'salary'])
                .select_related('branch', 'marked_by')
                .order_by('date')
            )
            
            for i, tr in enumerate(other_transactions, 1):
                ws3.append([
                    i,
                    tr.branch.name if tr.branch else "-",
                    tr.get_transaction_type_display(),
                    tr.get_category_display(),
                    tr.title,
                    float(tr.amount) if tr.amount else 0,
                    tr.marked_by.get_full_name() or tr.marked_by.username if tr.marked_by else "Tizim",
                    tr.date.strftime("%Y-%m-%d"),
                    tr.description or "-"
                ])
                self._apply_row_style(ws3[ws3.max_row], ws3)
            
            # Sheet 4: Summary Statistics
            ws4 = wb.create_sheet("Umumiy Statistika")
            headers4 = [
                "Filial Nomi", "O'quvchilar To'lovi", "Qo'shimcha Kirim",
                "Xodimlar Maoshi", "Refundlar (Chiqim)", "Kommunal To'lovlar",
                "Boshqa Chiqimlar", "SOF FOYDA", "Qarzdorlik"
            ]
            ws4.append(headers4)
            self._apply_header_style(ws4[1], ws4)
            
            branches = Branch.objects.all()
            grand_totals = {
                'income_fees': 0,
                'income_extra': 0,
                'expense_salary': 0,
                'expense_refund': 0,
                'expense_utility': 0,
                'expense_other': 0,
                'debt': 0
            }
            
            for branch in branches:
                # Calculate branch statistics
                income_fees = Payment.objects.filter(
                    group__branch=branch, month__year=year, month__month=month, is_paid=True
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                income_extra = FinanceTransaction.objects.filter(
                    branch=branch, date__year=year, date__month=month, 
                    transaction_type='income', category='student_extra'
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                expense_salary = EmployeePayment.objects.filter(
                    employee__branch=branch, month__year=year, month__month=month, is_paid=True
                ).aggregate(total=Sum(F('salary_base') + F('bonus') - F('deductions')))['total'] or 0
                
                expense_refund = FinanceTransaction.objects.filter(
                    branch=branch, date__year=year, date__month=month, 
                    transaction_type='expense', category='refund'
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                expense_utility = FinanceTransaction.objects.filter(
                    branch=branch, date__year=year, date__month=month, 
                    transaction_type='expense', category='utility'
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                expense_other = FinanceTransaction.objects.filter(
                    branch=branch, date__year=year, date__month=month, transaction_type='expense'
                ).exclude(category__in=['salary', 'refund', 'utility']).aggregate(total=Sum('amount'))['total'] or 0
                
                debt = Payment.objects.filter(
                    group__branch=branch, month__year=year, month__month=month, is_paid=False
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                net_profit = (income_fees + income_extra) - (expense_salary + expense_refund + expense_utility + expense_other)
                
                ws4.append([
                    branch.name,
                    float(income_fees) if income_fees else 0,
                    float(income_extra) if income_extra else 0,
                    float(expense_salary) if expense_salary else 0,
                    float(expense_refund) if expense_refund else 0,
                    float(expense_utility) if expense_utility else 0,
                    float(expense_other) if expense_other else 0,
                    float(net_profit) if net_profit else 0,
                    float(debt) if debt else 0
                ])
                self._apply_row_style(ws4[ws4.max_row], ws4)
                
                # Update grand totals
                grand_totals['income_fees'] += income_fees or 0
                grand_totals['income_extra'] += income_extra or 0
                grand_totals['expense_salary'] += expense_salary or 0
                grand_totals['expense_refund'] += expense_refund or 0
                grand_totals['expense_utility'] += expense_utility or 0
                grand_totals['expense_other'] += expense_other or 0
                grand_totals['debt'] += debt or 0
            
            # Grand total row
            grand_net = (grand_totals['income_fees'] + grand_totals['income_extra']) - \
                       (grand_totals['expense_salary'] + grand_totals['expense_refund'] + \
                        grand_totals['expense_utility'] + grand_totals['expense_other'])
            
            total_row = [
                "UMUMIY JAMI",
                float(grand_totals['income_fees']) if grand_totals['income_fees'] else 0,
                float(grand_totals['income_extra']) if grand_totals['income_extra'] else 0,
                float(grand_totals['expense_salary']) if grand_totals['expense_salary'] else 0,
                float(grand_totals['expense_refund']) if grand_totals['expense_refund'] else 0,
                float(grand_totals['expense_utility']) if grand_totals['expense_utility'] else 0,
                float(grand_totals['expense_other']) if grand_totals['expense_other'] else 0,
                float(grand_net) if grand_net else 0,
                float(grand_totals['debt']) if grand_totals['debt'] else 0
            ]
            ws4.append(total_row)
            for cell in ws4[ws4.max_row]:
                cell.font = Font(bold=True)
                cell.border = self.BORDER
            
            # Auto-adjust columns for all sheets
            for sheet in wb.worksheets:
                self._auto_adjust_columns(sheet)
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            self.logger.info(f"Successfully generated monthly financial report for {year}/{month}")
            return output
            
        except Exception as e:
            self.logger.error(f"Error generating monthly financial report: {str(e)}")
            raise ReportGenerationError(f"Failed to generate monthly financial report: {str(e)}")


class ReportDistributor:
    """
    Handles distribution of reports to authenticated Telegram users.
    """
    
    def __init__(self):
        self.logger = logger
        self.generator = ReportGenerator()
    
    def _send_file_to_chat(self, chat_id: str, file_buffer: io.BytesIO, filename: str, caption: str = ""):
        """
        Send a file to a Telegram chat.
        
        Args:
            chat_id: Telegram chat ID
            file_buffer: BytesIO object containing the file
            filename: Name of the file
            caption: Optional caption for the file
        """
        from django.conf import settings
        import requests
        
        # Use main bot token (not separate admin bot)
        bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', '')
        
        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
        
        file_buffer.seek(0)
        files = {
            'document': (filename, file_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        data = {
            'chat_id': chat_id,
            'caption': caption
        }
        
        try:
            response = requests.post(url, files=files, data=data, timeout=30)
            response.raise_for_status()
            self.logger.info(f"Successfully sent {filename} to chat {chat_id}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to send {filename} to chat {chat_id}: {str(e)}")
            return False
    
    def send_daily_report_to_admin(self, admin_user, target_date: Optional[date] = None):
        """
        Send daily report to an admin user.
        
        Args:
            admin_user: User object with admin role
            target_date: Date for the report (defaults to today)
        """
        if not admin_user.telegram_chat_id:
            self.logger.warning(f"Admin {admin_user.username} has no telegram_chat_id")
            return False
        
        if not admin_user.branch:
            self.logger.warning(f"Admin {admin_user.username} has no branch assigned")
            return False
        
        try:
            report_buffer = self.generator.generate_daily_branch_report(admin_user.branch, target_date)
            filename = f"daily_report_{target_date.strftime('%Y-%m-%d') if target_date else timezone.now().date().strftime('%Y-%m-%d')}.xlsx"
            caption = f"Kunlik hisobot - {admin_user.branch.name}"
            
            return self._send_file_to_chat(admin_user.telegram_chat_id, report_buffer, filename, caption)
            
        except ReportGenerationError as e:
            self.logger.error(f"Failed to send daily report to admin {admin_user.username}: {str(e)}")
            return False
    
    def send_monthly_attendance_to_admin(self, admin_user, year: int, month: int):
        """
        Send monthly attendance ZIP to an admin user.
        
        Args:
            admin_user: User object with admin role
            year: Year for the report
            month: Month for the report
        """
        if not admin_user.telegram_chat_id:
            self.logger.warning(f"Admin {admin_user.username} has no telegram_chat_id")
            return False
        
        if not admin_user.branch:
            self.logger.warning(f"Admin {admin_user.username} has no branch assigned")
            return False
        
        try:
            zip_buffer = self.generator.generate_monthly_attendance_zip(admin_user.branch, year, month)
            filename = f"attendance_{year}_{month:02d}.zip"
            caption = f"Oylik davomat arxivi - {admin_user.branch.name} ({year}-{month:02d})"
            
            return self._send_file_to_chat(admin_user.telegram_chat_id, zip_buffer, filename, caption)
            
        except ReportGenerationError as e:
            self.logger.error(f"Failed to send monthly attendance to admin {admin_user.username}: {str(e)}")
            return False
    
    def send_monthly_financial_to_super_admin(self, super_admin_user, year: int, month: int):
        """
        Send monthly financial report to a super_admin user.
        
        Args:
            super_admin_user: User object with super_admin role
            year: Year for the report
            month: Month for the report
        """
        if not super_admin_user.telegram_chat_id:
            self.logger.warning(f"Super Admin {super_admin_user.username} has no telegram_chat_id")
            return False
        
        try:
            report_buffer = self.generator.generate_monthly_financial_report(year, month)
            filename = f"financial_report_{year}_{month:02d}.xlsx"
            caption = f"Oylik moliyaviy hisobot - {year}-{month:02d}"
            
            return self._send_file_to_chat(super_admin_user.telegram_chat_id, report_buffer, filename, caption)
            
        except ReportGenerationError as e:
            self.logger.error(f"Failed to send monthly financial to super admin {super_admin_user.username}: {str(e)}")
            return False
