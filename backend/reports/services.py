import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
from django.db.models import Prefetch
from homework_attends.models import Attendance, Homework, HomeworkSubmission
from groups.models import Group, Student, GroupEnrollment

def generate_daily_full_report(target_date=None):
    if target_date is None:
        target_date = timezone.now().date()

    # Yangi Excel kitobi yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Hisobot {target_date.strftime('%Y.%m.%d')}"

    # --- Stillar ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Sarlavhalar ---
    headers = [
        "№", "Filial", "Guruh", "O'quvchi F.I.SH", 
        "Telefon", "Ota-ona telefon", "Davomat"
    ]
    ws.append(headers)

    # Sarlavha stilini qo'llash
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # --- Ma'lumotlarni yig'ish ---
    # 1. Faol guruhlardagi o'quvchilarni guruhlar bo'yicha saralangan holda olamiz
    enrollments = GroupEnrollment.objects.filter(
        is_active=True, 
        group__is_faol=True
    ).select_related(
        'student', 'group', 'group__branch'
    ).order_by('group__branch__name', 'group__name', 'student__full_name')

    # Bugungi davomatlarni keshga olamiz
    attendances = Attendance.objects.filter(date=target_date)
    att_dict = {(a.student_id, a.group_id): a for a in attendances}

    row_idx = 1
    reported_student_ids = set()

    # Guruhli o'quvchilarni chiqarish
    for enroll in enrollments:
        student = enroll.student
        group = enroll.group
        branch = group.branch
        reported_student_ids.add(student.id)
        
        is_lesson_day = group.is_lesson_day(target_date)
        att_status = "Belgilanmagan ⚠️"

        if not is_lesson_day:
            att_status = "Dars yo'q"
        else:
            att = att_dict.get((student.id, group.id))
            if att:
                att_status = "✅ Keldi" if att.is_present else "❌ Kelmadi"

        row_data = [
            row_idx,
            branch.name if branch else "Filialsiz",
            group.name,
            student.full_name,
            student.phone or "-",
            student.parent_phone or "-",
            att_status
        ]
        ws.append(row_data)

        # Stillar
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="left")
            if cell.column in [1, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_idx += 1

    # 2. Guruhsiz o'quvchilarni oxiriga qo'shmaymiz (talabga asosan olib tashlandi)

    # --- Ustun kengliklarini avtomatik sozlash ---
    dims = {
        'A': 5,  # №
        'B': 20, # Filial
        'C': 20, # Guruh
        'D': 35, # F.I.SH
        'E': 15, # Telefon
        'F': 15, # Ota-ona tel
        'G': 20  # Davomat
    }
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    return wb
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum, F
from finance.models import Payment, EmployeePayment, FinanceTransaction
from branches.models import Branch

def get_full_monthly_report(year, month):
    wb = openpyxl.Workbook()
    
    # --- Stillar ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # ================================================================
    # 1-SAHIFA: O'QUVCHILAR TO'LOVLARI (BATAFSIL RO'YXAT)
    # ================================================================
    ws1 = wb.active
    ws1.title = "O'quvchilar To'lovlari"
    
    headers1 = [
        "№", "Filial", "O'quvchi F.I.SH", "Guruh", "To'lov Summasi", 
        "To'lov Oyi", "Tasdiqlagan Admin", "To'langan Vaqt", "Tranzaksiya ID"
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    # Ma'lumotlarni olish
    student_payments = Payment.objects.filter(
        month__year=year, month__month=month, is_paid=True
    ).select_related('student', 'group', 'group__branch', 'marked_by').order_by('group__branch', 'student__full_name')

    for i, sp in enumerate(student_payments, 1):
        ws1.append([
            i,
            sp.group.branch.name if sp.group.branch else "-",
            sp.student.full_name,
            sp.group.name,
            sp.amount,
            sp.month.strftime("%Y-%m") if sp.month else "-",
            sp.marked_by.get_full_name() or sp.marked_by.username if sp.marked_by else "Onlayn",
            sp.paid_at.strftime("%Y-%m-%d %H:%M") if sp.paid_at else "-",
            sp.transaction_id or "-"
        ])

    # ================================================================
    # 2-SAHIFA: XODIMLAR MAOSHLARI (BATAFSIL RO'YXAT)
    # ================================================================
    ws2 = wb.create_sheet("Xodimlar Maoshlari")
    headers2 = [
        "№", "Filial", "Xodim F.I.SH", "Roli", "Asosiy Maosh", 
        "Bonus", "Jarimalar", "JAMI TO'LANDI", "Tasdiqladi (SuperAdmin)", "Sana"
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    employee_payments = EmployeePayment.objects.filter(
        month__year=year, month__month=month, is_paid=True
    ).select_related('employee', 'employee__branch', 'marked_by').order_by('employee__branch')

    for i, ep in enumerate(employee_payments, 1):
        total_paid = float(ep.salary_base + ep.bonus - ep.deductions)
        ws2.append([
            i,
            ep.employee.branch.name if ep.employee.branch else "-",
            ep.employee.get_full_name() or ep.employee.username,
            ep.employee.get_role_display(),
            ep.salary_base,
            ep.bonus,
            ep.deductions,
            total_paid,
            ep.marked_by.get_full_name() or ep.marked_by.username if ep.marked_by else "Tizim",
            ep.paid_at.strftime("%Y-%m-%d") if ep.paid_at else "-"
        ])

    # ================================================================
    # 3-SAHIFA: BOSHQA AMALLAR (REFUND, EXTRA, UTILITY)
    # ================================================================
    ws_extra = wb.create_sheet("Refund va Qo'shimcha Amallar")
    headers_extra = [
        "№", "Filial", "Tur", "Turkum", "Sarlavha", "Summa", "Mas'ul", "Sana", "Tavsif"
    ]
    ws_extra.append(headers_extra)
    for cell in ws_extra[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    other_transactions = FinanceTransaction.objects.filter(
        date__year=year, date__month=month
    ).exclude(category__in=['student_fee', 'salary']).select_related('branch', 'marked_by').order_by('date')

    for i, tr in enumerate(other_transactions, 1):
        ws_extra.append([
            i,
            tr.branch.name if tr.branch else "-",
            tr.get_transaction_type_display(),
            tr.get_category_display(),
            tr.title,
            tr.amount,
            tr.marked_by.get_full_name() or tr.marked_by.username if tr.marked_by else "Tizim",
            tr.date.strftime("%Y-%m-%d"),
            tr.description or "-"
        ])

    # ================================================================
    # 3.5-SAHIFA: ADMIN HARAJATLARI (MAYDA CHIQIMLAR)
    # ================================================================
    from finance.models import AdminExpense
    ws_admin = wb.create_sheet("Admin Harajatlari")
    headers_admin = [
        "№", "Filial", "Harajat Nomi", "Summa", "Kiritdi", "Sana", "Tavsif"
    ]
    ws_admin.append(headers_admin)
    for cell in ws_admin[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    admin_expenses = AdminExpense.objects.filter(
        date__year=year, date__month=month
    ).select_related('branch', 'marked_by').order_by('date')

    for i, ae in enumerate(admin_expenses, 1):
        ws_admin.append([
            i,
            ae.branch.name if ae.branch else "-",
            ae.title,
            ae.amount,
            ae.marked_by.get_full_name() or ae.marked_by.username if ae.marked_by else "Tizim",
            ae.date.strftime("%Y-%m-%d"),
            ae.description or "-"
        ])

    # ================================================================
    # 4-SAHIFA: FILIALLAR VA UMUMIY STATISTIKA (TUGATILDI)
    # ================================================================
    ws3 = wb.create_sheet("Umumiy Statistika")
    headers3 = [
        "Filial Nomi", "O'quvchilar To'lovi", "Qo'shimcha Kirim", 
        "Xodimlar Maoshlari", "Admin Harajatlari", "Refundlar (Chiqim)", "Kommunal To'lovlar",
        "Boshqa Chiqimlar", "SOF FOYDA", "Qarzdorlik"
    ]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    branches = Branch.objects.all()
    grand_income_fees = 0
    grand_income_extra = 0
    grand_expense_salary = 0
    grand_expense_admin = 0
    grand_expense_refund = 0
    grand_expense_utility = 0
    grand_expense_other = 0
    grand_debt = 0

    for branch in branches:
        # 1. Kirimlar
        income_fees = Payment.objects.filter(group__branch=branch, month__year=year, month__month=month, is_paid=True).aggregate(total=Sum('amount'))['total'] or 0
        income_extra = FinanceTransaction.objects.filter(branch=branch, date__year=year, date__month=month, transaction_type='income', category='student_extra').aggregate(total=Sum('amount'))['total'] or 0
        
        # 2. Chiqimlar
        expense_salary = EmployeePayment.objects.filter(employee__branch=branch, month__year=year, month__month=month, is_paid=True).aggregate(total=Sum(F('salary_base') + F('bonus') - F('deductions')))['total'] or 0
        expense_refund = FinanceTransaction.objects.filter(branch=branch, date__year=year, date__month=month, transaction_type='expense', category='refund').aggregate(total=Sum('amount'))['total'] or 0
        expense_utility = FinanceTransaction.objects.filter(branch=branch, date__year=year, date__month=month, transaction_type='expense', category='utility').aggregate(total=Sum('amount'))['total'] or 0
        expense_other = FinanceTransaction.objects.filter(branch=branch, date__year=year, date__month=month, transaction_type='expense').exclude(category__in=['salary', 'refund', 'utility']).aggregate(total=Sum('amount'))['total'] or 0
        
        # 3. Admin Harajatlari
        expense_admin = AdminExpense.objects.filter(branch=branch, date__year=year, date__month=month).aggregate(total=Sum('amount'))['total'] or 0
        
        # 4. Qarzdorlik
        debt = Payment.objects.filter(group__branch=branch, month__year=year, month__month=month, is_paid=False).aggregate(total=Sum('amount'))['total'] or 0
        
        net_profit = (income_fees + income_extra) - (expense_salary + expense_admin + expense_refund + expense_utility + expense_other)
        
        ws3.append([
            branch.name, 
            income_fees, 
            income_extra, 
            expense_salary, 
            expense_admin,
            expense_refund, 
            expense_utility,
            expense_other, 
            net_profit, 
            debt
        ])
        
        grand_income_fees += income_fees
        grand_income_extra += income_extra
        grand_expense_salary += expense_salary
        grand_expense_admin += expense_admin
        grand_expense_refund += expense_refund
        grand_expense_utility += expense_utility
        grand_expense_other += expense_other
        grand_debt += debt

    # Jami qatori
    total_net = (grand_income_fees + grand_income_extra) - (grand_expense_salary + grand_expense_admin + grand_expense_refund + grand_expense_utility + grand_expense_other)
    total_row = [
        "UMUMIY JAMI", 
        grand_income_fees, 
        grand_income_extra, 
        grand_expense_salary, 
        grand_expense_admin,
        grand_expense_refund, 
        grand_expense_utility,
        grand_expense_other, 
        total_net, 
        grand_debt
    ]
    ws3.append(total_row)
    for cell in ws3[ws3.max_row]:
        cell.font = Font(bold=True)
        cell.border = border

    # --- Ustun kengliklarini matn uzunligiga qarab avtomatik to'g'rilash ---
    for sheet in wb.worksheets:
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value:
                        val_len = len(str(cell.value))
                        if val_len > max_length: max_length = val_len
                except: pass
            sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 50)

    return wb