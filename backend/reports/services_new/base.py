import openpyxl
from django.utils import timezone
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class BaseReportService:
    def __init__(self, user, target_date=None, is_manual=False):
        self.user = user
        self.target_date = target_date or timezone.now().date()
        self.is_manual = is_manual
        self.wb = openpyxl.Workbook()

        # Common styles
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    def apply_security_scope(self, queryset, branch_field='group__branch'):
        """Enforces branch-level data isolation."""
        if not self.user:
            return queryset
            
        if self.user.role == 'super_admin':
            return queryset
        elif self.user.role == 'admin':
            if not self.user.branch:
                return queryset.none()
            filter_kwargs = {branch_field: self.user.branch}
            return queryset.filter(**filter_kwargs)
        else:
            return queryset.none()

    def apply_headers(self, ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.border

    def auto_adjust_widths(self, ws, dims):
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    def generate(self):
        """Must be implemented by subclasses to build the workbook."""
        raise NotImplementedError
