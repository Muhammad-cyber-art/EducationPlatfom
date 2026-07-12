from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
from .base import BaseReportService
from groups.models import Group
from homework_attends.services import get_monthly_attendance_data
import gc

class GroupAttendanceReport(BaseReportService):
    def __init__(self, user, group_id, target_year, target_month, is_manual=False):
        super().__init__(user, is_manual=is_manual)
        self.group_id = group_id
        self.target_year = target_year
        self.target_month = target_month
        self.group = Group.objects.get(id=group_id)

    def generate(self):
        # Security: ensure the user has access to this group's branch
        if self.user.role == 'admin':
            allowed = [self.user.branch_id] if self.user.branch_id else []
            if hasattr(self.user, 'branch_accesses'):
                allowed.extend(self.user.branch_accesses.values_list('branch_id', flat=True))
            if self.group.branch_id not in allowed:
                raise PermissionError("Sizda ushbu guruhga kirish huquqi yo'q")

        ws = self.wb.active
        
        safe_sheet_name = f"Davomat_{self.group.name[:20]}"
        for char in r"\/?:*[]":
            safe_sheet_name = safe_sheet_name.replace(char, "")
        safe_sheet_name = safe_sheet_name.strip()[:30]
        if not safe_sheet_name:
            safe_sheet_name = "Davomat"
        ws.title = safe_sheet_name
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        date_list, students, att_data = get_monthly_attendance_data(self.group, self.target_month, self.target_year)
        
        headers = ["№", "O'quvchi ismi-familiyasi", "Telefon", "Ota-ona telefon"] + [d.day for d in date_list]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        for col_num in range(5, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 4
            
        header_row = ws[2]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = self.border

        for row_num, student in enumerate(students, 3):
            ws.cell(row=row_num, column=1).value = row_num - 2
            ws.cell(row=row_num, column=2).value = student.full_name
            ws.cell(row=row_num, column=3).value = student.phone or "-"
            ws.cell(row=row_num, column=4).value = student.parent_phone or "-"
            
            joined_at = student.joined_at.date() if student.joined_at else None
            
            for col_idx, d in enumerate(date_list, 5):
                cell = ws.cell(row=row_num, column=col_idx)
                
                att_record = att_data.get(student.id, {}).get(str(d))
                
                if att_record:
                    if not att_record.get('is_confirmed'):
                        cell.value = "!" 
                        cell.font = Font(color="FF0000", bold=True)
                    elif att_record.get('is_present') is True:
                        cell.value = "+"
                        cell.font = Font(color="008000", bold=True)
                    else:
                        cell.value = "K" 
                        cell.font = Font(color="FF0000", bold=True)
                elif joined_at and d < joined_at:
                    cell.value = "-"
                else:
                    cell.value = "!"
                    cell.font = Font(color="FF0000", bold=True)

                cell.alignment = center_align
                cell.border = self.border

            for col_num in range(1, 5):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(vertical="center", horizontal="left") if col_num in [2,3,4] else center_align
                cell.border = self.border

        gc.collect()
        return self.wb
