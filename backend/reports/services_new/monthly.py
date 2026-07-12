from openpyxl.styles import Font, Alignment
import gc
from openpyxl.utils import get_column_letter
from django.db.models import Sum, F
from django.utils import timezone
from .base import BaseReportService
from finance.models import Payment, EmployeePayment, FinanceTransaction
from branches.models import Branch
from homework_attends.models import Attendance, Homework, HomeworkSubmission

class MonthlyAttendanceReport(BaseReportService):
    def __init__(self, user, target_year, target_month, is_manual=False):
        super().__init__(user, is_manual=is_manual)
        self.target_year = target_year
        self.target_month = target_month

    def generate(self):
        ws = self.wb.active
        ws.title = f"Davomat {self.target_year}-{self.target_month:02d}"

        headers = [
            "№", "Sana", "Filial", "Guruh", "O'quvchi F.I.SH", 
            "Davomat", "Uy vazifasi (Mavzu)", "Vazifa holati"
        ]
        self.apply_headers(ws, headers)

        # Build base queryset
        qs = Attendance.objects.filter(
            date__year=self.target_year,
            date__month=self.target_month,
            group__isnull=False,
            student__isnull=False,
        )
        if self.is_manual:
            qs = qs.filter(date__lte=timezone.now().date())
            
        qs = qs.select_related('student', 'group', 'group__branch').order_by('date', 'group__branch__name', 'group__name', 'student__full_name')
        
        # Apply security
        qs = self.apply_security_scope(qs, branch_field='group__branch')

        # Since it's for a whole month, we won't prefetch all homeworks for memory efficiency,
        # but we can cache homeworks lightly or just use a simple lookup since it's a generated report.
        # For simplicity and performance, we use a basic dictionary cache for homeworks.
        homework_cache = {}

        for index, att in enumerate(qs.iterator(chunk_size=1000), start=1):
            student = att.student
            group = att.group
            branch = group.branch

            cache_key = f"{group.id}_{att.date}"
            if cache_key not in homework_cache:
                hw = Homework.objects.filter(group=group, created_at__date=att.date).first()
                homework_cache[cache_key] = hw
                
            homework = homework_cache[cache_key]
            hw_title = "Vazifa berilmagan"
            hw_status = "-"

            if homework:
                hw_title = homework.title
                submission = HomeworkSubmission.objects.filter(homework=homework, student=student).first()
                hw_status = submission.get_status_display() if submission else "Topshirmagan ❌"

            row_data = [
                index,
                att.date.strftime('%Y-%m-%d'),
                branch.name if branch else "Filialsiz",
                group.name,
                student.full_name,
                "✅ Keldi" if att.is_present else "❌ Kelmadi",
                hw_title,
                hw_status
            ]
            ws.append(row_data)

            for cell in ws[ws.max_row]:
                cell.border = self.border
                cell.alignment = Alignment(vertical="center", horizontal="left")
                if cell.column in [1, 2, 6, 8]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        dims = {'A': 5, 'B': 15, 'C': 20, 'D': 20, 'E': 35, 'F': 15, 'G': 30, 'H': 20}
        self.auto_adjust_widths(ws, dims)
        return self.wb


class MonthlyFinanceReport(BaseReportService):
    def __init__(self, user, target_year, target_month, is_manual=False):
        super().__init__(user, is_manual=is_manual)
        self.target_year = target_year
        self.target_month = target_month

    def _auto_adjust_all_sheets(self):
        for sheet in self.wb.worksheets:
            for col_idx, column_cells in enumerate(sheet.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        if cell.value:
                            val_len = len(str(cell.value))
                            if val_len > max_length: max_length = val_len
                    except: pass
                sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 50)

    def generate(self):
        align_center = Alignment(horizontal="center", vertical="center")

        # ================================================================
        # 1. O'QUVCHILAR TO'LOVLARI
        # ================================================================
        ws1 = self.wb.active
        ws1.title = "O'quvchilar To'lovlari"
        headers1 = ["№", "Filial", "O'quvchi F.I.SH", "Guruh", "To'lov Summasi", "To'lov Oyi", "Tasdiqlagan Admin", "To'langan Vaqt", "Tranzaksiya ID"]
        self.apply_headers(ws1, headers1)

        qs1 = Payment.objects.filter(
            month__year=self.target_year, month__month=self.target_month, is_paid=True,
            group__isnull=False, group__branch__isnull=False, student__isnull=False
        )
        if self.is_manual:
            qs1 = qs1.filter(paid_at__date__lte=timezone.now().date())
            
        qs1 = self.apply_security_scope(qs1, branch_field='group__branch')
        qs1 = qs1.select_related('student', 'group', 'group__branch', 'marked_by').order_by('group__branch', 'student__full_name')

        for i, sp in enumerate(qs1.iterator(chunk_size=1000), 1):
            ws1.append([
                i, sp.group.branch.name if sp.group.branch else "-", sp.student.full_name, sp.group.name,
                sp.amount, sp.month.strftime("%Y-%m") if sp.month else "-",
                sp.marked_by.get_full_name() or sp.marked_by.username if sp.marked_by else "Onlayn",
                sp.paid_at.strftime("%Y-%m-%d %H:%M") if sp.paid_at else "-", sp.transaction_id or "-"
            ])
            for cell in ws1[ws1.max_row]:
                cell.border = self.border
                cell.alignment = align_center

        # ================================================================
        # 2. XODIMLAR MAOSHLARI
        # ================================================================
        ws2 = self.wb.create_sheet("Xodimlar Maoshlari")
        headers2 = ["№", "Filial", "Xodim F.I.SH", "Roli", "Asosiy Maosh", "Bonus", "Jarimalar", "JAMI TO'LANDI", "Tasdiqladi", "Sana"]
        self.apply_headers(ws2, headers2)

        qs2 = EmployeePayment.objects.filter(month__year=self.target_year, month__month=self.target_month, is_paid=True)
        if self.is_manual:
            qs2 = qs2.filter(paid_at__date__lte=timezone.now().date())
            
        qs2 = self.apply_security_scope(qs2, branch_field='employee__branch')
        qs2 = qs2.select_related('employee', 'employee__branch', 'marked_by').order_by('employee__branch')

        for i, ep in enumerate(qs2.iterator(chunk_size=1000), 1):
            total_paid = float(ep.salary_base + ep.bonus - ep.deductions)
            ws2.append([
                i, ep.employee.branch.name if ep.employee.branch else "-", ep.employee.get_full_name() or ep.employee.username,
                ep.employee.get_role_display(), ep.salary_base, ep.bonus, ep.deductions, total_paid,
                ep.marked_by.get_full_name() or ep.marked_by.username if ep.marked_by else "Tizim",
                ep.paid_at.strftime("%Y-%m-%d") if ep.paid_at else "-"
            ])
            for cell in ws2[ws2.max_row]:
                cell.border = self.border
                cell.alignment = align_center

        # ================================================================
        # 3. BOSHQA AMALLAR
        # ================================================================
        ws_extra = self.wb.create_sheet("Boshqa Amallar")
        headers_extra = ["№", "Filial", "Tur", "Turkum", "Sarlavha", "Summa", "Mas'ul", "Sana", "Tavsif"]
        self.apply_headers(ws_extra, headers_extra)

        qs3 = FinanceTransaction.objects.filter(date__year=self.target_year, date__month=self.target_month).exclude(category__in=['student_fee', 'salary'])
        if self.is_manual:
            qs3 = qs3.filter(date__lte=timezone.now().date())
            
        qs3 = self.apply_security_scope(qs3, branch_field='branch')
        qs3 = qs3.select_related('branch', 'marked_by').order_by('date')

        for i, tr in enumerate(qs3.iterator(chunk_size=1000), 1):
            ws_extra.append([
                i, tr.branch.name if tr.branch else "-", tr.get_transaction_type_display(), tr.get_category_display(),
                tr.title, tr.amount, tr.marked_by.get_full_name() or tr.marked_by.username if tr.marked_by else "Tizim",
                tr.date.strftime("%Y-%m-%d"), tr.description or "-"
            ])
            for cell in ws_extra[ws_extra.max_row]:
                cell.border = self.border
                cell.alignment = align_center

        # ================================================================
        # 4. UMUMIY STATISTIKA (Only for SuperAdmin or Branch specific)
        # ================================================================
        ws3 = self.wb.create_sheet("Umumiy Statistika")
        headers3 = ["Filial Nomi", "O'quvchilar To'lovi", "Qo'shimcha Kirim", "Xodimlar Maoshi", "Refundlar (Chiqim)", "Kommunal", "Boshqa Chiqimlar", "SOF FOYDA", "Qarzdorlik"]
        self.apply_headers(ws3, headers3)

        branches = Branch.objects.all()
        if self.user and self.user.role == 'admin':
            branches = branches.filter(id=self.user.branch_id) if self.user.branch_id else branches.none()

        grand = {k: 0 for k in ['inc_fees', 'inc_extra', 'exp_salary', 'exp_refund', 'exp_util', 'exp_other', 'debt']}

        for branch in branches:
            # Queries inside loop since branches are few. For optimization, aggregation could be grouped, but this matches original logic.
            inc_fees = Payment.objects.filter(group__branch=branch, month__year=self.target_year, month__month=self.target_month, is_paid=True).aggregate(t=Sum('amount'))['t'] or 0
            inc_extra = FinanceTransaction.objects.filter(branch=branch, date__year=self.target_year, date__month=self.target_month, transaction_type='income', category='student_extra').aggregate(t=Sum('amount'))['t'] or 0
            
            exp_salary = EmployeePayment.objects.filter(employee__branch=branch, month__year=self.target_year, month__month=self.target_month, is_paid=True).aggregate(t=Sum(F('salary_base') + F('bonus') - F('deductions')))['t'] or 0
            exp_refund = FinanceTransaction.objects.filter(branch=branch, date__year=self.target_year, date__month=self.target_month, transaction_type='expense', category='refund').aggregate(t=Sum('amount'))['t'] or 0
            exp_util = FinanceTransaction.objects.filter(branch=branch, date__year=self.target_year, date__month=self.target_month, transaction_type='expense', category='utility').aggregate(t=Sum('amount'))['t'] or 0
            exp_other = FinanceTransaction.objects.filter(branch=branch, date__year=self.target_year, date__month=self.target_month, transaction_type='expense').exclude(category__in=['salary', 'refund', 'utility']).aggregate(t=Sum('amount'))['t'] or 0
            
            debt = Payment.objects.filter(group__branch=branch, month__year=self.target_year, month__month=self.target_month, is_paid=False).aggregate(t=Sum('amount'))['t'] or 0
            
            net_profit = (inc_fees + inc_extra) - (exp_salary + exp_refund + exp_util + exp_other)
            
            ws3.append([branch.name, inc_fees, inc_extra, exp_salary, exp_refund, exp_util, exp_other, net_profit, debt])
            
            grand['inc_fees'] += inc_fees
            grand['inc_extra'] += inc_extra
            grand['exp_salary'] += exp_salary
            grand['exp_refund'] += exp_refund
            grand['exp_util'] += exp_util
            grand['exp_other'] += exp_other
            grand['debt'] += debt
            
            for cell in ws3[ws3.max_row]:
                cell.border = self.border
                cell.alignment = align_center

        total_net = (grand['inc_fees'] + grand['inc_extra']) - (grand['exp_salary'] + grand['exp_refund'] + grand['exp_util'] + grand['exp_other'])
        total_row = ["UMUMIY JAMI", grand['inc_fees'], grand['inc_extra'], grand['exp_salary'], grand['exp_refund'], grand['exp_util'], grand['exp_other'], total_net, grand['debt']]
        ws3.append(total_row)
        for cell in ws3[ws3.max_row]:
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = align_center

        self._auto_adjust_all_sheets()
        
        gc.collect()
        return self.wb
